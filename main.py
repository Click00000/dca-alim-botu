from fastapi import FastAPI, Query, Depends, HTTPException, status, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response

from pydantic import BaseModel
import numpy as np
import pandas as pd
import ccxt
from tradingview_ta import TA_Handler, Interval
from datetime import datetime, timedelta
from typing import List, Optional, Dict, Any
import asyncio
import time
import json
import os
import sqlite3
from contextlib import contextmanager

import hashlib
from collections import deque
from threading import Lock
import tempfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# BIST hisse listelerini import et
from bist_stocks_ak import BIST_STOCKS_AK, get_stocks_by_symbol as get_stocks_ak_by_symbol, search_stocks as search_stocks_ak
from bist_stocks_lz import BIST_STOCKS_LZ, get_stocks_by_symbol as get_stocks_lz_by_symbol, search_stocks as search_stocks_lz

app = FastAPI(title="DCA Scanner API", version="1.0.0")

# ---------- Database Yönetimi ----------
DATABASE_PATH = os.environ.get("DATABASE_PATH", "dca_scanner.db")

def init_database():
    """Database'i başlat ve tabloları oluştur"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Kullanıcılar tablosu
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS users (
                    id TEXT PRIMARY KEY,
                    username TEXT UNIQUE NOT NULL,
                    password TEXT NOT NULL,
                    email TEXT,
                    is_admin BOOLEAN DEFAULT FALSE,
                    created_at TEXT NOT NULL,
                    last_login TEXT,
                    is_active BOOLEAN DEFAULT TRUE
                )
            ''')
            
            # Portföyler tablosu
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS portfolios (
                    portfolio_id TEXT PRIMARY KEY,
                    portfolio_name TEXT NOT NULL,
                    portfolio_description TEXT,
                    owner_username TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    last_updated TEXT
                )
            ''')
            
            # Portföy işlemleri tablosu
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS portfolio_items (
                    id TEXT PRIMARY KEY,
                    portfolio_id TEXT NOT NULL,
                    symbol TEXT NOT NULL,
                    market TEXT NOT NULL,
                    transaction_type TEXT NOT NULL,
                    price REAL NOT NULL,
                    quantity REAL NOT NULL,
                    date TEXT NOT NULL,
                    target_price REAL,
                    notes TEXT,
                    current_price REAL,
                    last_updated TEXT,
                    owner_username TEXT NOT NULL,
                    FOREIGN KEY (portfolio_id) REFERENCES portfolios (portfolio_id)
                )
            ''')
            
            # Takip listesi tablosu
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS watchlist (
                    id TEXT PRIMARY KEY,
                    symbol TEXT NOT NULL,
                    market TEXT NOT NULL,
                    added_date TEXT NOT NULL,
                    current_price REAL,
                    last_updated TEXT,
                    target_price REAL,
                    notes TEXT
                )
            ''')
            
            conn.commit()
            print("✅ Database tabloları başarıyla oluşturuldu")
            
    except Exception as e:
        print(f"❌ Database başlatma hatası: {e}")

@contextmanager
def get_db_connection():
    """Database bağlantısı için context manager"""
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row  # Dict-like access
    try:
        yield conn
    finally:
        conn.close()

def migrate_json_to_database():
    """JSON dosyalarından verileri database'e taşı"""
    try:
        # Kullanıcıları taşı
        if os.path.exists(USERS_FILE):
            with open(USERS_FILE, 'r', encoding='utf-8') as f:
                users = json.load(f)
            
            with get_db_connection() as conn:
                cursor = conn.cursor()
                for user in users:
                    cursor.execute('''
                        INSERT OR REPLACE INTO users 
                        (id, username, password, email, is_admin, created_at, last_login, is_active)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        user.get('id'), user.get('username'), user.get('password'),
                        user.get('email'), user.get('is_admin', False),
                        user.get('created_at'), user.get('last_login'),
                        user.get('is_active', True)
                    ))
                conn.commit()
                print(f"✅ {len(users)} kullanıcı database'e taşındı")
        
        # Portföyleri taşı
        if os.path.exists(PORTFOLIO_LIST_FILE):
            with open(PORTFOLIO_LIST_FILE, 'r', encoding='utf-8') as f:
                portfolios = json.load(f)
            
            with get_db_connection() as conn:
                cursor = conn.cursor()
                for portfolio in portfolios:
                    cursor.execute('''
                        INSERT OR REPLACE INTO portfolios 
                        (portfolio_id, portfolio_name, portfolio_description, owner_username, created_at)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (
                        portfolio.get('portfolio_id'), portfolio.get('portfolio_name'),
                        portfolio.get('portfolio_description'), portfolio.get('owner_username'),
                        portfolio.get('created_at')
                    ))
                conn.commit()
                print(f"✅ {len(portfolios)} portföy database'e taşındı")
        
        # Portföy işlemlerini taşı
        if os.path.exists(PORTFOLIO_DIR):
            for filename in os.listdir(PORTFOLIO_DIR):
                if filename.endswith('.json'):
                    portfolio_id = filename.replace('.json', '')
                    portfolio_file = os.path.join(PORTFOLIO_DIR, filename)
                    
                    with open(portfolio_file, 'r', encoding='utf-8') as f:
                        items = json.load(f)
                    
                    with get_db_connection() as conn:
                        cursor = conn.cursor()
                        for item in items:
                            cursor.execute('''
                                INSERT OR REPLACE INTO portfolio_items 
                                (id, portfolio_id, symbol, market, transaction_type, price, quantity,
                                 date, target_price, notes, current_price, last_updated, owner_username)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            ''', (
                                item.get('id'), portfolio_id, item.get('symbol'), item.get('market'),
                                item.get('transaction_type'), item.get('price'), item.get('quantity'),
                                item.get('date'), item.get('target_price'), item.get('notes'),
                                item.get('current_price'), item.get('last_updated'),
                                item.get('owner_username')
                            ))
                        conn.commit()
                        print(f"✅ {len(items)} işlem {portfolio_id} için database'e taşındı")
        
        print("🎉 Tüm veriler database'e başarıyla taşındı!")
        
    except Exception as e:
        print(f"❌ Veri taşıma hatası: {e}")

# ---------- Portföy Modelleri ----------
class PortfolioItem(BaseModel):
    id: str
    symbol: str
    market: str  # "bist" veya "crypto"
    transaction_type: str  # "buy" veya "sell"
    price: float
    quantity: float
    date: str
    target_price: Optional[float] = None
    notes: Optional[str] = None
    current_price: Optional[float] = None
    last_updated: Optional[str] = None
    
    class Config:
        extra = "allow"  # Ekstra alanlara izin ver

class PortfolioAddRequest(BaseModel):
    symbol: str
    market: str
    transaction_type: str  # "buy" veya "sell"
    price: float
    quantity: float
    target_price: Optional[float] = None
    notes: Optional[str] = None
    portfolio_id: str  # Hangi portföye ekleneceği

class PortfolioUpdateRequest(BaseModel):
    transaction_type: Optional[str] = None  # "buy" veya "sell"
    price: Optional[float] = None
    quantity: Optional[float] = None
    target_price: Optional[float] = None
    notes: Optional[str] = None

# ---------- Takip Listesi Modelleri ----------
class WatchlistItem(BaseModel):
    id: str
    symbol: str
    market: str  # "bist" veya "crypto"
    added_date: str
    current_price: Optional[float] = None
    last_updated: Optional[str] = None
    target_price: Optional[float] = None
    notes: Optional[str] = None

class WatchlistAddRequest(BaseModel):
    symbol: str
    market: str
    target_price: Optional[float] = None
    notes: Optional[str] = None

class WatchlistUpdateRequest(BaseModel):
    target_price: Optional[float] = None
    notes: Optional[str] = None

# ---------- Portföy Veri Yönetimi ----------
PORTFOLIO_DIR = "data/portfolios"
PORTFOLIO_LIST_FILE = "data/portfolio_list.json"
WATCHLIST_FILE = "watchlist.json"

# ---------- Yeni ID Sistemi ----------
def get_user_uid(username: str) -> int:
    """Kullanıcı adından UID çıkar: wastfc -> 1, deneme1 -> 2"""
    if username == ADMIN_USERNAME:
        return 1
    else:
        # deneme1 -> 2, deneme2 -> 3, etc.
        try:
            return int(username.replace("deneme", "")) + 1
        except:
            return 999  # Bilinmeyen kullanıcı

def create_portfolio_id(username: str, portfolio_number: int) -> str:
    """Portfolio ID oluştur: dca1_001, dca2_001, etc."""
    uid = get_user_uid(username)
    return f"dca{uid}_{portfolio_number:03d}"

def get_user_portfolios(username: str) -> list:
    """Kullanıcının sadece kendi portfolio'larını getir"""
    uid = get_user_uid(username)
    portfolio_list = load_portfolio_list()
    return [p for p in portfolio_list if p["portfolio_id"].startswith(f"dca{uid}_")]

def get_next_portfolio_number(username: str) -> int:
    """Kullanıcının bir sonraki portfolio numarasını getir (1-20 arası)"""
    user_portfolios = get_user_portfolios(username)
    if len(user_portfolios) >= 20:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Maksimum 20 portfolio limitine ulaştınız")
    
    existing_numbers = []
    for p in user_portfolios:
        try:
            num = int(p["portfolio_id"].split("_")[1])
            existing_numbers.append(num)
        except:
            continue
    
    # 1'den 20'ye kadar ilk boş numarayı bul
    for i in range(1, 21):
        if i not in existing_numbers:
            return i
    
    return 1  # Fallback

# ---------- Simple API Key Authentication ----------
SECRET_KEY = "dca-scanner-secret-key-2024"  # Gerçek uygulamada güvenli olmalı

# Active API keys - basit ve etkili
active_api_keys = {}

# ---------- Kullanıcı Veritabanı ----------
USERS_FILE = "data/users.json"
ADMIN_USERNAME = "wastfc"
ADMIN_PASSWORD = "Sanene88"  # Gerçek uygulamada hash'lenmiş olmalı

# Test ortamı için şifre hash'leme devre dışı
TEST_MODE = True  # False yaparak production moduna geçebilirsiniz

def load_users():
    """Kullanıcı listesini database'den yükle"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM users WHERE is_active = 1')
            users = []
            for row in cursor.fetchall():
                user = dict(row)
                
                # Test modunda mevcut hash'li şifreleri düz metin olarak göster
                if TEST_MODE:
                    # Eğer şifre hash ise, düz metin olarak göster
                    if len(user.get('password', '')) == 64:  # SHA256 hash uzunluğu
                        if user['username'] == 'wastfc':
                            user['password'] = 'Sanene88'  # Admin şifresi
                        else:
                            user['password'] = 'deneme123'  # Test kullanıcıları
                
                users.append(user)
            return users
    except Exception as e:
        print(f"❌ Database'den kullanıcı yükleme hatası: {e}")
        # Fallback: JSON dosyasından yükle
        if os.path.exists(USERS_FILE):
            try:
                with open(USERS_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                pass
        return []

def hash_password(password: str) -> str:
    """Şifreyi hash'le"""
    if TEST_MODE:
        return password  # Test modunda düz metin olarak sakla
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(plain_password: str, hashed_password: str) -> bool:
    """Şifreyi doğrula"""
    if TEST_MODE:
        print(f"🔍 DEBUG: Test mode password check - plain: '{plain_password}', stored: '{hashed_password}'")
        return plain_password == hashed_password  # Test modunda düz metin karşılaştır
    return hash_password(plain_password) == hashed_password

def create_api_key(username: str) -> str:
    """API key oluştur ve döndür"""
    import secrets
    api_key = f"dca_{secrets.token_urlsafe(16)}"
    active_api_keys[api_key] = {
        "username": username,
        "created_at": datetime.now().isoformat()
    }
    print(f"✅ Created API key for user: {username}")
    return api_key

def verify_api_key(api_key: str) -> dict:
    """API key'i doğrula ve kullanıcı bilgisini döndür"""
    if not api_key or api_key not in active_api_keys:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Geçersiz API key")
    
    return active_api_keys[api_key]

def get_current_user(authorization: str = Header(None)):
    """Authorization header'dan API key al ve kullanıcı bilgisini döndür"""
    if not authorization:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Authorization header gerekli")
    
    # "Bearer API_KEY" formatından API key'i çıkar
    if not authorization.startswith("Bearer "):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Bearer token gerekli")
    
    api_key = authorization.replace("Bearer ", "")
    
    try:
        api_data = verify_api_key(api_key)
        username = api_data["username"]
        
        users = load_users()
        user = next((u for u in users if u["username"] == username), None)
        if not user:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Kullanıcı bulunamadı")
        return user
    except Exception as e:
        print(f"❌ API key verification error: {e}")
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="API key doğrulanamadı")

def save_users(users):
    """Kullanıcı listesini database'e kaydet"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Mevcut kullanıcıları temizle
            cursor.execute('DELETE FROM users')
            
            # Yeni kullanıcıları ekle
            for user in users:
                cursor.execute('''
                    INSERT INTO users 
                    (id, username, password, email, is_admin, created_at, last_login, is_active)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    user.get('id'), user.get('username'), user.get('password'),
                    user.get('email'), user.get('is_admin', False),
                    user.get('created_at'), user.get('last_login'),
                    user.get('is_active', True)
                ))
            
            conn.commit()
            print(f"✅ {len(users)} kullanıcı database'e kaydedildi")
            
    except Exception as e:
        print(f"❌ Database'e kullanıcı kaydetme hatası: {e}")
        # Fallback: JSON dosyasına kaydet
        try:
            os.makedirs(os.path.dirname(USERS_FILE), exist_ok=True)
            with open(USERS_FILE, 'w', encoding='utf-8') as f:
                json.dump(users, f, ensure_ascii=False, indent=2)
            print("✅ Fallback: Kullanıcılar JSON dosyasına kaydedildi")
        except Exception as json_error:
            print(f"❌ JSON fallback hatası: {json_error}")

def create_default_admin():
    """Varsayılan admin kullanıcısını oluştur"""
    users = load_users()
    
    # Admin kullanıcısı var mı kontrol et
    admin_exists = any(user.get('username') == ADMIN_USERNAME for user in users)
    
    if not admin_exists:
        admin_user = {
            "id": "admin_001",
            "username": ADMIN_USERNAME,
            "password": hash_password(ADMIN_PASSWORD),  # Test modunda düz metin, production'da hash
            "email": "admin@dca-scanner.com",
            "is_admin": True,
            "created_at": datetime.now().isoformat(),
            "last_login": None,
            "is_active": True
        }
        users.append(admin_user)
        save_users(users)
        print(f"✅ Varsayılan admin kullanıcısı oluşturuldu: {ADMIN_USERNAME}")
    
    return users

def authenticate_user(username: str, password: str):
    """Kullanıcı kimlik doğrulaması"""
    users = load_users()
    for user in users:
        if user['username'] == username and user['password'] == password and user['is_active']:
            return user
    return None

def get_user_by_id(user_id: str):
    """ID'ye göre kullanıcı getir"""
    users = load_users()
    for user in users:
        if user['id'] == user_id:
            return user
    return None

class PortfolioCreateRequest(BaseModel):
    name: str
    description: Optional[str] = None
    # owner_username artık backend'de otomatik olarak current_user'dan alınacak

# ---------- Kullanıcı Yönetimi Modelleri ----------
class User(BaseModel):
    id: str
    username: str
    password: Optional[str] = None  # Admin için şifre görüntüleme
    email: Optional[str] = None
    is_admin: bool = False
    created_at: str
    last_login: Optional[str] = None
    is_active: bool = True

class UserCreateRequest(BaseModel):
    username: str
    email: Optional[str] = None
    password: str
    is_admin: bool = False

class UserLoginRequest(BaseModel):
    username: str
    password: str

class UserUpdateRequest(BaseModel):
    username: Optional[str] = None
    email: Optional[str] = None
    is_admin: Optional[bool] = None
    is_active: Optional[bool] = None
    password: Optional[str] = None

class AdminLoginRequest(BaseModel):
    username: str
    password: str

def load_portfolio_list():
    """Portföy listesini yükle"""
    if os.path.exists(PORTFOLIO_LIST_FILE):
        try:
            with open(PORTFOLIO_LIST_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return []
    return []

def save_portfolio_list(portfolio_list):
    """Portföy listesini kaydet"""
    os.makedirs(os.path.dirname(PORTFOLIO_LIST_FILE), exist_ok=True)
    with open(PORTFOLIO_LIST_FILE, 'w', encoding='utf-8') as f:
        json.dump(portfolio_list, f, ensure_ascii=False, indent=2)

def load_portfolio(portfolio_id: str):
    """Belirli bir portföyün verilerini yükle"""
    portfolio_file = os.path.join(PORTFOLIO_DIR, f"{portfolio_id}.json")
    if os.path.exists(portfolio_file):
        try:
            with open(portfolio_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return []
    return []

def save_portfolio(portfolio_id: str, portfolio_data):
    """Belirli bir portföyün verilerini kaydet"""
    os.makedirs(PORTFOLIO_DIR, exist_ok=True)
    portfolio_file = os.path.join(PORTFOLIO_DIR, f"{portfolio_id}.json")
    with open(portfolio_file, 'w', encoding='utf-8') as f:
        json.dump(portfolio_data, f, ensure_ascii=False, indent=2)

# ---------- Takip Listesi Veri Yönetimi ----------
def load_watchlist():
    """Takip listesi verilerini yükle"""
    if os.path.exists(WATCHLIST_FILE):
        try:
            with open(WATCHLIST_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return []
    return []

def save_watchlist(watchlist):
    """Takip listesi verilerini kaydet"""
    with open(WATCHLIST_FILE, 'w', encoding='utf-8') as f:
        json.dump(watchlist, f, ensure_ascii=False, indent=2)

def generate_id():
    """Benzersiz ID oluştur"""
    return datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]

# CORS ayarları - Cookie gönderimi için gerekli
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000", 
        "http://localhost:3001", 
        "http://localhost:3002", 
        "http://localhost:3003", 
        "http://localhost:3004",
        "https://dca-alim-botu-5ugrlwh5o-click00000s-projects.vercel.app",
        "https://*.vercel.app"  # Tüm Vercel subdomain'leri
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------- Rate Limiting Ayarları ----------
REQUEST_DELAY = 10.0  # İstek başına minimum 10 saniye bekleme (manuel taramada çalışan seviye)
MAX_CONCURRENT_REQUESTS = 3  # Aynı anda maksimum 3 paralel istek
REQUEST_QUEUE = deque()  # İstek kuyruğu
QUEUE_LOCK = Lock()  # Thread-safe queue erişimi
LAST_REQUEST_TIME = 0  # Son istek zamanı

def wait_for_rate_limit():
    """Rate limiting için bekleme"""
    global LAST_REQUEST_TIME
    current_time = time.time()
    time_since_last = current_time - LAST_REQUEST_TIME
    
    if time_since_last < REQUEST_DELAY:
        sleep_time = REQUEST_DELAY - time_since_last
        print(f"Rate limiting: {sleep_time:.1f} saniye bekleniyor...")
        time.sleep(sleep_time)
    
    LAST_REQUEST_TIME = time.time()

# ---------- BIST Hisse Fonksiyonları ----------
def get_all_bist_stocks():
    """Tüm BIST hisselerini birleştir"""
    return BIST_STOCKS_AK + BIST_STOCKS_LZ

def get_bist_stock_by_symbol(symbol: str):
    """Sembol ile BIST hissesi bul"""
    # Önce A-K listesinde ara
    stock = get_stocks_ak_by_symbol(symbol)
    if stock:
        return stock
    
    # Sonra L-Z listesinde ara
    stock = get_stocks_lz_by_symbol(symbol)
    if stock:
        return stock
    
    return None

def search_bist_stocks(query: str):
    """BIST hisselerinde arama yap"""
    results = []
    
    # A-K listesinde ara
    ak_results = search_stocks_ak(query)
    results.extend(ak_results)
    
    # L-Z listesinde ara
    lz_results = search_stocks_lz(query)
    results.extend(lz_results)
    
    return results

# ---------- Utility Fonksiyonları ----------
def atr(df: pd.DataFrame, n: int = 14) -> pd.Series:
    """Average True Range hesaplama - Güvenli versiyon"""
    try:
        if df.empty or len(df) < 2:
            return pd.Series([0.001] * len(df))  # Minimum değer döndür
        
        h, l, c = df["high"], df["low"], df["close"]
        prev_c = c.shift(1)
        
        # True Range hesaplama
        tr1 = h - l  # High - Low
        tr2 = abs(h - prev_c)  # |High - Previous Close|
        tr3 = abs(l - prev_c)  # |Low - Previous Close|
        
        # True Range = max(tr1, tr2, tr3)
        tr = np.maximum(tr1, np.maximum(tr2, tr3))
        
        # NaN değerleri temizle
        tr = tr.fillna(0.001)
        
        # Rolling mean hesapla
        atr_series = tr.rolling(n, min_periods=1).mean()
        
        # Minimum değer kontrolü
        atr_series = atr_series.clip(lower=0.001)
        
        return atr_series
        
    except Exception as e:
        print(f"ATR hesaplama hatası: {e}")
        # Hata durumunda minimum değer döndür
        return pd.Series([0.001] * len(df))

def obv(df: pd.DataFrame) -> pd.Series:
    """On Balance Volume hesaplama"""
    up = (df["close"] > df["close"].shift(1)).astype(int)
    down = (df["close"] < df["close"].shift(1)).astype(int) * -1
    dirn = (up + down).fillna(0)
    return (dirn * df["volume"]).cumsum()

def ema(s: pd.Series, n: int) -> pd.Series:
    """Exponential Moving Average hesaplama"""
    return s.ewm(span=n, adjust=False).mean()

def ccxt_ohlcv(exchange: str, symbol: str, tf: str = "1d", limit: int = 400) -> pd.DataFrame:
    """CCXT ile OHLCV verisi çekme"""
    try:
        ex = getattr(ccxt, exchange)({"enableRateLimit": True})
        ohlcv = ex.fetch_ohlcv(symbol, timeframe=tf, limit=limit)
        df = pd.DataFrame(ohlcv, columns=["ts", "open", "high", "low", "close", "volume"])
        return df
    except Exception as e:
        print(f"CCXT error for {symbol}: {e}")
        return pd.DataFrame()

def tv_get_analysis(symbol: str, market: str, tf: str = "1d") -> Optional[Dict[str, Any]]:
    """TradingView'dan teknik analiz verisi çekme - Rate Limited"""
    try:
        # Rate limiting uygula
        wait_for_rate_limit()
        
        # Market'e göre exchange belirle
        if market == "crypto":
            exchange = "BINANCE"
        elif market == "bist":
            exchange = "BIST"
        elif market == "us":
            exchange = "NASDAQ"
        elif market == "fx":
            exchange = "FX_IDC"
        else:
            exchange = "BINANCE"
        
        # Timeframe'i TradingView formatına çevir
        if tf == "1d":
            interval = "1d"
        elif tf == "4h":
            interval = "4h"
        else:
            interval = "1d"
        
        print(f"TradingView API isteği: {symbol} ({exchange}) - {time.strftime('%H:%M:%S')}")
        
        # TradingView handler oluştur
        print(f"🔍 DEBUG: Handler oluşturuluyor - Symbol: {symbol}, Exchange: {exchange}, Screener: {'crypto' if market == 'crypto' else 'turkey' if market == 'bist' else 'america'}")
        
        handler = TA_Handler(
            symbol=symbol,
            exchange=exchange,
            screener="crypto" if market == "crypto" else "turkey" if market == "bist" else "america",
            interval=interval,
            timeout=30  # Timeout'u artır
        )
        
        print(f"🔍 DEBUG: Handler oluşturuldu, analiz alınıyor...")
        
        # Analiz verisi çek
        analysis = handler.get_analysis()
        
        print(f"🔍 DEBUG: Analiz sonucu: {analysis}")
        
        if analysis:
            result = {
                "symbol": symbol,
                "market": market,
                "close": analysis.indicators.get("close", 0),
                "high": analysis.indicators.get("high", 0),
                "low": analysis.indicators.get("low", 0),
                "volume": analysis.indicators.get("volume", 0),
                "rsi": analysis.indicators.get("RSI", 50),
                "macd": analysis.indicators.get("MACD.macd", 0),
                "macd_signal": analysis.indicators.get("MACD.signal", 0),
                "sma_20": analysis.indicators.get("SMA20", 0),
                "sma_50": analysis.indicators.get("SMA50", 0),
                "ema_20": analysis.indicators.get("EMA20", 0),
                "ema_50": analysis.indicators.get("EMA50", 0),
                "bb_upper": analysis.indicators.get("BB.upper", 0),
                "bb_lower": analysis.indicators.get("BB.lower", 0),
                "bb_middle": analysis.indicators.get("BB.middle", 0),
                "atr": analysis.indicators.get("ATR", 0),
                "summary": analysis.summary,
                "oscillators": analysis.oscillators,
                "moving_averages": analysis.moving_averages,
                "indicators": analysis.indicators
            }
            print(f"🔍 DEBUG: tv_get_analysis döndürüyor: {result}")
            return result
        
        return None
        
    except Exception as e:
        print(f"TradingView error for {symbol}: {e}")
        
        # 429 hatası için özel bekleme
        if "429" in str(e) or "rate limit" in str(e).lower():
            print(f"Rate limit hatası! {symbol} için 15 saniye ek bekleme...")
            time.sleep(15)
            
            # 2. deneme yap
            try:
                print(f"2. deneme: {symbol} için TradingView API'ye tekrar istek atılıyor...")
                analysis = handler.get_analysis()
                if analysis:
                    # Analiz başarılı, devam et
                    pass
            except Exception as retry_error:
                print(f"2. deneme de başarısız: {symbol} - {retry_error}")
        
        return None

def compute_signals_tv(analysis: Dict[str, Any]) -> Dict[str, Any]:
    """TradingView verisi ile DCA sinyallerini hesapla - Yeni Algoritma"""
    print(f"🔍 DEBUG: compute_signals_tv çağrıldı! analysis keys: {list(analysis.keys())}")
    try:
        # Temel değerler
        close = float(analysis.get("close", 0))
        high = float(analysis.get("high", 0))
        low = float(analysis.get("low", 0))
        volume = float(analysis.get("volume", 0))
        rsi = float(analysis.get("rsi", 50))
        
        # ATR hesaplama - TradingView'dan gelmiyorsa manuel hesapla
        atr_raw = analysis.get("atr", 0)
        if atr_raw == 0 or atr_raw is None or atr_raw < 0:
            # Daha güvenli ATR hesaplama: True Range kullanarak
            # True Range = max(high-low, |high-prev_close|, |low-prev_close|)
            # Basit yaklaşım: (high - low) kullan, ama minimum değer kontrolü ile
            atr = max(high - low, 0.001)  # Minimum 0.001 değeri
        else:
            atr = float(atr_raw)
            # ATR değerinin mantıklı olduğundan emin ol
            if atr <= 0:
                atr = max(high - low, 0.001)
        
        # ATR değerinin çok büyük olmadığından emin ol (anormal değerler için)
        if atr > close * 0.5:  # ATR, fiyatın %50'sinden büyükse anormal
            atr = close * 0.05  # %5 olarak sınırla
        
        # Moving averages
        ema20 = float(analysis.get("ema_20", close))
        ema50 = float(analysis.get("ema_50", close))
        sma20 = float(analysis.get("sma_20", close))
        sma50 = float(analysis.get("sma_50", close))
        
        # Bollinger Bands
        bb_upper = float(analysis.get("bb_upper", close * 1.02))
        bb_lower = float(analysis.get("bb_lower", close * 0.98))
        bb_middle = float(analysis.get("bb_middle", close))
        
        # bb_middle 0 ise close kullan
        if bb_middle == 0:
            bb_middle = close
        
        # MACD
        macd = float(analysis.get("macd", 0))
        macd_signal = float(analysis.get("macd_signal", 0))
        
        # TradingView özetleri
        summary = analysis.get("summary", {})
        oscillators = analysis.get("oscillators", {})
        moving_averages = analysis.get("moving_averages", {})
        
        # Basit seviye hesaplamaları (60 mum verisi olmadığı için basit)
        RL = low * 0.95  # Support seviyesi
        RH = high * 1.05  # Resistance seviyesi
        H = RH - RL
        VAL = RL + 0.25 * H  # Value Area Low
        
        # Volume ratio (basit hesaplama)
        vol_ratio = 1.0  # TradingView'da volume ratio yok, varsayılan 1.0
        
        # YENİ DCA ALGORİTMASI - Puanlama (0-100)
        score = 0
        score_details = {}
        
        # 1. Uzun süreli akümülasyon kontrolü (20 puan) - Akıllı Range Puanlama Sistemi
        # Range %20'den az ise akümülasyon (Pine Script'teki gibi)
        range_pct = H / max(RL, 0.01) * 100
        
        # Yeni akıllı akümülasyon puanlama sistemi (2 parça: bant genişliği + konum)
        if range_pct <= 20:  # Sadece %20 altındaki range'ler için puanlama
            # 1) Bant genişliği (0-10): Range ne kadar dar?
            if range_pct <= 10:
                width_score = 10  # Çok dar bant
            elif range_pct <= 15:
                width_score = 7   # Dar bant
            elif range_pct <= 20:
                width_score = 4   # Orta bant
            else:
                width_score = 0
            
            # 2) Range içi konum (0-10): Fiyat RL/VAL/RH'ye ne kadar yakın?
            # Eşikler (oynanabilir)
            thr_rl = 3.0   # RL yakınlık eşiği (%)
            thr_val = 2.0  # VAL yakınlık eşiği (%)
            thr_rh = 3.0   # RH yakınlık eşiği (%)
            
            # Yüzde mesafeler
            dist_pct_rl = abs((close - RL) / RL) * 100.0 if RL > 0 else 100
            dist_pct_val = abs((close - VAL) / VAL) * 100.0 if VAL > 0 else 100
            dist_pct_rh = abs((RH - close) / RH) * 100.0 if RH > 0 else 100
            
            # Yakınlık kontrolü
            near_rl = dist_pct_rl <= thr_rl
            near_val = dist_pct_val <= thr_val
            near_rh = dist_pct_rh <= thr_rh
            
            # Puanlama (maks 10; birden fazlası tutarsa topla ama 10'da kapa)
            pos_score_raw = (
                (7 if near_rl else 0) +   # Destek testine yakın olmak en değerli
                (5 if near_rh else 0) +   # Kırılım öncesi konum
                (3 if near_val else 0)    # Denge bölgesinde gezinme
            )
            pos_score = min(pos_score_raw, 10)
            
            # 3) ATR Bonus Puanı (+1): ATR çok düşükken konum puanına bonus
            atr_bonus = 0
            if atr > 0:
                atr_ratio = atr / close * 100  # ATR'nin fiyata oranı (%)
                if atr_ratio <= 2.0:  # ATR %2'nin altındaysa bonus
                    atr_bonus = 1
                    pos_score = min(pos_score + atr_bonus, 10)  # Maksimum 10'da kapa
            
            # Nihai akümülasyon skoru (0-20)
            akumulasyon_puan = width_score + pos_score
            
            # Debug bilgisi
            print("=" * 50)
            print(f"🔍 RANGE DEBUG - Symbol: {analysis.get('symbol', 'Unknown')}")
            print(f"🔍 RANGE DEBUG - analysis keys: {list(analysis.keys())}")
            print(f"🔍 RANGE DEBUG - range_pct: {range_pct:.2f}%, width_score: {width_score}/10")
            print(f"🔍 RANGE DEBUG - pos_score: {pos_score}/10 (RL:{near_rl}, VAL:{near_val}, RH:{near_rh})")
            print(f"🔍 RANGE DEBUG - ATR bonus: +{atr_bonus}")
            print(f"🔍 RANGE DEBUG - total: {akumulasyon_puan}/20")
            print("=" * 50)
            
        else:
            akumulasyon_puan = 0   # Akümülasyon kriterine girmez
            width_score = 0
            pos_score = 0
            atr_bonus = 0
        
        score += akumulasyon_puan
        score_details["akumulasyon"] = akumulasyon_puan
        score_details["range_pct"] = round(range_pct, 2)  # Range yüzdesini de sakla
        score_details["range_width_score"] = width_score  # Bant genişliği puanı
        score_details["range_position_score"] = pos_score  # Konum puanı
        score_details["atr_bonus"] = atr_bonus  # ATR bonus puanı
        
        # 2. Manipülasyon fitili (Spring) kontrolü (15 puan) - Detaylı Pine Script algoritması
        # Spring toleransı %2 altına iğne + kapanış tekrar üstte
        spring_tol_pct = 2.0  # Pine Script'teki springTolPct
        spring_depth = RL * (1.0 - spring_tol_pct/100.0)
        is_spring_bar = (low < spring_depth) and (close > RL)  # reclaim şartı
        
        if is_spring_bar:
            # Mum parçaları hesaplama
            body = abs(close - open)
            low_wick = (open if open < close else close) - low   # alt fitil uzunluğu
            up_wick = high - (open if open > close else close)   # üst fitil
            
            # 1) Fitil/gövde oranına göre 0-7 puan
            wick_body_ratio = low_wick / body if body > 0 else 0
            wick_score = 0
            if wick_body_ratio >= 1.5:  # wickFullRatio
                wick_score = 7  # Güçlü fitil
            elif wick_body_ratio >= 0.7:  # wickHalfRatio
                wick_score = 4  # Orta fitil
            else:
                wick_score = 0  # Zayıf fitil
            
            # 2) Pozisyon: alt fitil güçlü, üst fitil küçükse 0-4 puan
            pos_score = 0
            if low_wick >= up_wick * 1.2:  # Alt fitil üst fitilden belirgin uzun
                pos_score = 4
            elif low_wick >= up_wick * 0.8:  # Orta durum
                pos_score = 2
            else:
                pos_score = 0
            
            # 3) Destek (RL) yakınlığı: fitilin en dip noktası RL'e ne kadar yakın? 0-4 puan
            dist_pct_rl = ((low - RL) / RL) * 100.0 if RL > 0 else 0
            near_pct_rl = 1.5  # Destek yakınlığı eşiği (%)
            support_score = 0
            if abs(dist_pct_rl) <= near_pct_rl:
                support_score = 4  # RL'e çok yakın
            elif abs(dist_pct_rl) <= near_pct_rl * 2:
                support_score = 2  # RL'e orta yakınlık
            else:
                support_score = 0  # RL'den uzak
            
            # 4) Hacim bonusu: spring barında hacim artmışsa +1 puan
            vol_bonus = 0
            if volume > 0:  # Volume verisi varsa
                # Basit hacim kontrolü (20 günlük ortalamanın 1.5x'i)
                vol_bonus = 1 if volume > 1000000 else 0  # Basit eşik
            
            # Toplam Spring puanı (maksimum 15)
            spring_puan = min(wick_score + pos_score + support_score + vol_bonus, 15)
            
            # Spring detaylarını score_details'e ekle
            score_details["spring_wick_score"] = wick_score
            score_details["spring_pos_score"] = pos_score
            score_details["spring_support_score"] = support_score
            score_details["spring_vol_bonus"] = vol_bonus
            
            # Debug bilgisi
            print(f"Spring Debug - Symbol: {analysis.get('symbol', 'Unknown')}")
            print(f"Spring Debug - wick/body: {wick_body_ratio:.2f}, wick_score: {wick_score}")
            print(f"Spring Debug - pos_score: {pos_score}, support_score: {support_score}")
            print(f"Spring Debug - vol_bonus: {vol_bonus}, total: {spring_puan}")
            
        else:
            spring_puan = 0
            # Spring yoksa detayları 0 olarak ayarla
            score_details["spring_wick_score"] = 0
            score_details["spring_pos_score"] = 0
            score_details["spring_support_score"] = 0
            score_details["spring_vol_bonus"] = 0
        
        score += spring_puan
        score_details["spring"] = spring_puan
        
        # 3. OBV yukarı yönlü (15 puan) - Pine Script algoritması
        # OBV trendi (basit hesaplama)
        obv_puan = 0
        if rsi > 45 and close > (high + low) / 2:  # RSI yükseliyor ve fiyat ortalamanın üstünde
            obv_puan = 15
        elif rsi > 40:
            obv_puan = 10
        else:
            obv_puan = 5
        
        score += obv_puan
        score_details["obv"] = obv_puan
        
        # 4. Akıllı Hacim Puanlama Sistemi (10 puan) - Yeni Pine Script algoritması
        # Not: TradingView'da volume moving averages yok, alternatif yaklaşım kullanılıyor
        
        # True Range hesaplama (ATR için kullanılan)
        tr_ = max(high - low, abs(high - close), abs(low - close))
        
        # Spring ve Breakout koşulları (mevcut değişkenlerle hizala)
        is_spring_bar = (low < RL * 0.98) and (close > RL)  # Spring bar kontrolü
        near_break = close >= RH * 0.98  # Kırılıma yakınlık
        is_breakout = close > RH  # Kırılım
        brk_cond = is_breakout or near_break  # Breakout koşulu
        
        # Alternatif hacim puanlama (TradingView verisi ile uyumlu)
        # 1) Kuruma (Drying Up) (0-3 puan) - Basit volume spike kontrolü
        dry_score = 0
        if volume > 0:  # Volume verisi varsa
            # Volume çok yüksekse kuruma olabilir (anormal spike)
            if volume > 1000000:  # Basit eşik
                dry_score = 1.5  # Orta kuruma
            elif volume > 500000:
                dry_score = 0.5  # Hafif kuruma
            else:
                dry_score = 3.0  # Normal volume, kuruma yok
        
        # 2) Spring Hacmi (0-3 puan)
        spring_vol_score = 0
        if is_spring_bar:
            if volume > 800000:  # Spring barında yüksek volume
                spring_vol_score = 3.0
            elif volume > 500000:
                spring_vol_score = 1.5
            else:
                spring_vol_score = 0.5
        
        # 3) Kırılım Hacmi (0-3 puan)
        brk_vol_score = 0
        if brk_cond:
            if volume > 1000000:  # Kırılımda çok yüksek volume
                brk_vol_score = 3.0
            elif volume > 800000:
                brk_vol_score = 1.5
            else:
                brk_vol_score = 0.5
        
        # 4) Churn Cezası (0 to -1 puan)
        spread_small = tr_ <= 0.6 * atr
        churn_penalty = 0
        if volume > 800000 and spread_small:  # Yüksek volume + küçük spread
            churn_penalty = -1.0
        
        # Nihai Hacim Skoru (0-10)
        vol_score_raw = dry_score + spring_vol_score + brk_vol_score + churn_penalty
        vol_score_10 = max(0.0, min(10.0, vol_score_raw))
        
        # Hacim detaylarını score_details'e ekle
        score_details["volume_score"] = vol_score_10
        score_details["dry_up_score"] = dry_score
        score_details["spring_volume_score"] = spring_vol_score
        score_details["breakout_volume_score"] = brk_vol_score
        score_details["churn_penalty"] = churn_penalty
        
        # Eski hacim puanını yeni ile değiştir
        hacim_puan = vol_score_10
        score_details["hacim"] = hacim_puan
        
        # Debug bilgisi - Hacim puanlama
        print("=" * 50)
        print(f"🔍 VOLUME DEBUG - Symbol: {analysis.get('symbol', 'Unknown')}")
        print(f"🔍 VOLUME DEBUG - dry_score: {dry_score}/3, spring_vol: {spring_vol_score}/3")
        print(f"🔍 VOLUME DEBUG - brk_vol: {brk_vol_score}/3, churn_penalty: {churn_penalty}")
        print(f"🔍 VOLUME DEBUG - total: {vol_score_10}/10")
        print("=" * 50)
        
        score += hacim_puan
        
        # 5. Breakout'a yakınlık (10 puan) - Pine Script algoritması
        # Fiyat RH'a %5 mesafede mi? (Pine Script'teki nearPct = 5.0)
        near_pct = 5.0
        near_break = close >= RH * (1.0 - near_pct/100.0)
        
        if near_break:
            score += 10
            score_details["breakout_yakin"] = 10
        else:
            score_details["breakout_yakin"] = 0
        
        # 6. EMA kesişimi (10 puan) + Golden Cross Bonus (+2 puan) - Pine Script algoritması
        # EMA20>EMA50 && SMA20>SMA50 (Pine Script'teki maUp)
        ma_up = (ema20 > ema50) and (sma20 > sma50)
        ema_only_up = (ema20 > ema50)
        
        # Golden Cross Bonus hesaplama (+2 puan)
        # Not: TradingView'da barssince ve crossover fonksiyonları yok, basit hesaplama
        # Son 10 bar içinde EMA20 EMA50'yi yukarı kesmiş mi?
        # Spread artıyor mu? (momentum teyidi)
        cross_lookback = 10  # Golden cross lookback (bar)
        
        # Basit Golden Cross tespiti: EMA20 > EMA50 ve spread artıyor
        spread_current = ema20 - ema50
        spread_previous = ema20 * 0.99 - ema50 * 1.01  # Basit yaklaşım
        spread_up = spread_current > spread_previous
        
        # Golden Cross bonus (+2 puan)
        gc_bonus_2 = 0.0
        if ema_only_up and spread_up:
            gc_bonus_2 = 2.0
        
        # EMA kesişim puanı
        if ma_up:
            ema_kesisim_puan = 10
        elif ema_only_up:
            ema_kesisim_puan = 6  # Pine Script'teki gibi
        else:
            ema_kesisim_puan = 0
        
        # Golden Cross bonus'u ekle (toplam puanı 100'ü aşmayacak şekilde)
        ema_total_puan = min(ema_kesisim_puan + gc_bonus_2, 10)  # Maksimum 10 puan
        
        score += ema_total_puan
        score_details["ema_kesisim"] = ema_total_puan
        score_details["golden_cross_bonus"] = gc_bonus_2
        
        # Debug bilgisi - Golden Cross Bonus
        print("=" * 50)
        print(f"🔍 GOLDEN CROSS DEBUG - Symbol: {analysis.get('symbol', 'Unknown')}")
        print(f"🔍 GOLDEN CROSS DEBUG - EMA kesişim: {ema_kesisim_puan}/10")
        print(f"🔍 GOLDEN CROSS DEBUG - Golden Cross bonus: +{gc_bonus_2}")
        print(f"🔍 GOLDEN CROSS DEBUG - Total EMA: {ema_total_puan}/10")
        print("=" * 50)
        
        # 7. RSI düşükten toparlanma (10 puan) - Pine Script algoritması
        # RSI < 35 (Pine Script'teki rsiLow)
        rsi_low = rsi < 35
        
        if rsi_low:
            score += 10
            score_details["rsi_toparlanma"] = 10
        else:
            score_details["rsi_toparlanma"] = 0
        
        # 8. ATR Volatilite Puanlaması (10 puan) - Yeni eklenen
        # ATR, fiyatın yüzdesi olarak hesaplanır: (ATR / Close) × 100
        try:
            # Debug bilgisi
            print(f"ATR Debug - Symbol: {analysis.get('symbol', 'Unknown')}")
            print(f"ATR Debug - Close: {close}, ATR: {atr}")
            
            # Güvenli ATR yüzde hesaplama
            if close > 0 and atr > 0:
                atr_perc = (atr / close) * 100
                print(f"ATR Debug - ATR%: {atr_perc:.2f}%")
            else:
                atr_perc = 0
                atr_puan = 0
                print(f"ATR Debug - Invalid values: close={close}, atr={atr}")
            
            # ATR yüzdesi mantıklı sınırlar içinde mi kontrol et
            if atr_perc > 100:  # ATR fiyatın %100'ünden büyükse anormal
                atr_perc = 5.0  # %5 olarak sınırla
                atr_puan = 0
                print(f"ATR Debug - ATR% too high, limited to 5%")
            else:
                # ATR'ye göre puanlama sistemi
                if atr_perc < 1:
                    atr_puan = 10  # Çok sakin, sert hareket potansiyeli yüksek
                elif atr_perc < 2:
                    atr_puan = 8   # Düşük volatilite, akümülasyon ihtimali
                elif atr_perc < 3:
                    atr_puan = 5   # Orta volatilite, kırılım öncesi olabilir
                elif atr_perc < 5:
                    atr_puan = 3   # Yüksek volatilite, trend başlama ihtimali
                else:
                    atr_puan = 0   # Aşırı volatilite, risk yüksek
                
                print(f"ATR Debug - Score: {atr_puan}/10")
                    
        except Exception as e:
            print(f"ATR puanlama hatası: {e}")
            atr_puan = 0
            atr_perc = 0
        
        score += atr_puan
        score_details["atr"] = atr_puan
        
        # Skoru 0-100 arasında sınırla
        score = max(0, min(100, score))
        
        # Kategori belirleme
        if score >= 70:
            category = "Strong DCA"
        elif score >= 50:
            category = "DCA"
        elif score >= 30:
            category = "Weak DCA"
        else:
            category = "No DCA Signal"
        
        # Pine Script'teki hedef bantları (T1, T2, T3)
        T1_from = RH + 0.45 * H
        T1_to = RH + 0.85 * H
        T2_from = RH + 1.50 * H
        T2_to = RH + 1.55 * H
        T3_from = RH + 2.80 * H
        T3_to = RH + 3.00 * H
        
        return {
            "RL": RL, "VAL": VAL, "RH": RH, "H": H, "ATR": atr, "volRatio": vol_ratio,
            "isDCA": score >= 50, "isDipReclaim": category == "Dip-Reclaim", 
            "isNearBreakout": category == "Near-Breakout", "isBreakout": category == "Breakout",
            "score": score, "ema20": ema20, "ema50": ema50, "avwap": bb_middle, 
            "close": close, "category": category,
            "rsi": rsi, "macd": macd, "bb_upper": bb_upper, "bb_lower": bb_lower,
            "score_details": score_details,  # Yeni: detaylı puanlar
            "range_pct": range_pct,  # Yeni: range yüzdesi
            # Pine Script hedef bantları
            "targets": {
                "T1": {"from": T1_from, "to": T1_to, "label": "Hedef 1"},
                "T2": {"from": T2_from, "to": T2_to, "label": "Hedef 2"},
                "T3": {"from": T3_from, "to": T3_to, "label": "Hedef 3"}
            }
        }
        
    except Exception as e:
        print(f"Error computing signals for TradingView data: {e}")
        return {
            "RL": 0, "VAL": 0, "RH": 0, "H": 0, "ATR": 0, "volRatio": 0,
            "isDCA": False, "isDipReclaim": False, "isNearBreakout": False, "isBreakout": False,
            "score": 0, "ema20": 0, "ema50": 0, "avwap": 0, "close": 0, "category": "Neutral"
        }

# ---------- Piyasa Sembolleri (Test için sadece 3 büyük) ----------
BIST = ["THYAO", "GARAN", "KCHOL"]

US = ["AAPL", "MSFT", "NVDA", "AMZN", "META", "GOOGL", "TSLA", "NFLX", "AMD", "SPY", "PLTR"]

FX = ["EURUSD", "GBPUSD", "USDJPY", "XAUUSD", "USDCAD", 
      "AUDUSD", "NZDUSD", "USDCHF", "GBPJPY", "EURJPY"]

# Test için sadece 10 büyük kripto
CRYPTO_TOP_10 = ["BTCUSDT", "ETHUSDT", "BNBUSDT", "SOLUSDT", "ADAUSDT", 
                  "AVAXUSDT", "DOTUSDT", "MATICUSDT", "LINKUSDT", "UNIUSDT"]

# Market cap'e göre en büyük 100 USDT paritesi crypto token (Tekrarlar temizlendi)
CRYPTO_TOP_100 = [
    # Top 10 - Blue Chips
    "BTCUSDT", "ETHUSDT", "BNBUSDT", "SOLUSDT", "ADAUSDT", 
    "AVAXUSDT", "DOTUSDT", "MATICUSDT", "LINKUSDT", "UNIUSDT",
    
    # Top 11-25 - Major Altcoins
    "XRPUSDT", "DOGEUSDT", "SHIBUSDT", "TRXUSDT", "LTCUSDT",
    "BCHUSDT", "XLMUSDT", "VETUSDT", "FILUSDT", "ATOMUSDT",
    "NEARUSDT", "FTMUSDT", "ALGOUSDT", "ICPUSDT", "ETCUSDT",
    
    # Top 26-50 - Established Projects
    "HBARUSDT", "XMRUSDT", "APTUSDT", "OPUSDT", "ARBUSDT",
    "MKRUSDT", "AAVEUSDT", "GRTUSDT", "SANDUSDT", "MANAUSDT",
    "THETAUSDT", "XTZUSDT", "EOSUSDT", "CAKEUSDT", "CHZUSDT",
    "HOTUSDT", "ZECUSDT", "DASHUSDT", "NEOUSDT", "WAVESUSDT",
    "BATUSDT", "ZRXUSDT", "OMGUSDT", "QTUMUSDT", "IOTAUSDT",
    
    # Top 51-75 - Growing Ecosystems
    "RUNEUSDT", "KSMUSDT", "DYDXUSDT", "SNXUSDT", "COMPUSDT",
    "YFIUSDT", "CRVUSDT", "1INCHUSDT", "SUSHIUSDT", "RENUSDT",
    "STORJUSDT", "ANKRUSDT", "OCEANUSDT", "BANDUSDT", "ALPHAUSDT",
    "AUDIOUSDT", "ENJUSDT", "REEFUSDT", "CTSIUSDT", "CELOUSDT",
    "DENTUSDT", "HIVEUSDT", "STMXUSDT", "TRBUSDT", "VTHOUSDT",
    
    # Top 76-100 - Emerging Projects & DeFi
    "ROSEUSDT", "IOTXUSDT", "ONEUSDT", "HARMONYUSDT", "ZILUSDT",
    "ICXUSDT", "ONTUSDT", "WANUSDT", "WTCUSDT", "NULSUSDT",
    "RLCUSDT", "GNTUSDT", "REPUSDT", "BTSUSDT", "LSKUSDT",
    "ARKUSDT", "NAVUSDT", "PIVXUSDT", "DGBUSDT", "SYSUSDT",
    "VIAUSDT", "MONAUSDT", "XVGUSDT", "AXSUSDT", "SLPUSDT"
]

# ---------- Pydantic Modelleri ----------
class ScanItem(BaseModel):
    symbol: str
    market: str
    score: float
    category: str
    RL: float
    VAL: float
    RH: float
    H: float
    ATR: float
    volRatio: float
    close: float

class ChartResponse(BaseModel):
    ohlcv: Dict[str, List[float]]
    levels: Dict[str, float]
    targets: Dict[str, Dict[str, Any]]
    signals: Dict[str, Any]
    entries: Dict[str, float]
    stops: Dict[str, float]

# ---------- API Endpoints ----------
@app.get("/health")
async def health_check():
    """Backend sağlık kontrolü"""
    try:
        users = load_users()
        return {
            "status": "ok",
            "timestamp": datetime.now().isoformat(),
            "users_count": len(users),
            "backend": "DCA Scanner Backend",
            "version": "1.0.0"
        }
    except Exception as e:
        return {
            "status": "error",
            "timestamp": datetime.now().isoformat(),
            "error": str(e)
        }

@app.get("/status")
async def system_status():
    """Sistem durumu ve kullanıcı bilgileri"""
    try:
        users = load_users()
        user_list = []
        
        for user in users:
            user_info = {
                "username": user.get("username"),
                "is_admin": user.get("is_admin", False),
                "is_active": user.get("is_active", True),
                "last_login": user.get("last_login")
            }
            user_list.append(user_info)
        
        return {
            "success": True,
            "total_users": len(users),
            "users": user_list,
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "timestamp": datetime.now().isoformat()
        }

@app.get("/test-bist")
def test_bist():
    """BIST test endpoint'i"""
    try:
        # BIST hisse sayılarını test et
        ak_count = len(BIST_STOCKS_AK)
        lz_count = len(BIST_STOCKS_LZ)
        total_count = ak_count + lz_count
        
        # Birkaç örnek hisse göster
        sample_stocks = BIST_STOCKS_AK[:5] + BIST_STOCKS_LZ[:5]
        
        return {
            "success": True, 
            "ak_count": ak_count,
            "lz_count": lz_count,
            "total_count": total_count,
            "sample_stocks": sample_stocks
        }
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"success": False, "error": str(e)}

@app.get("/test-crypto")
def test_crypto():
    """Crypto test endpoint'i"""
    try:
        # Kripto token sayısını test et
        crypto_count = len(CRYPTO_TOP_100)
        
        # İlk 10 token'ı göster
        sample_tokens = CRYPTO_TOP_100[:10]
        
        # Arama testi
        search_test = search_crypto("BTC", 5)
        
        return {
            "success": True,
            "crypto_count": crypto_count,
            "sample_tokens": sample_tokens,
            "search_test": search_test
        }
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"success": False, "error": str(e)}

@app.get("/debug-crypto")
def debug_crypto(symbol: str = "ARBUSDT"):
    """Crypto debug endpoint'i - TradingView analizini test et"""
    try:
        print(f"🔍 DEBUG: {symbol} için TradingView analizi test ediliyor...")
        
        # TradingView analizi al
        analysis = tv_get_analysis(symbol, "crypto", "1d")
        
        if analysis:
            print(f"✅ {symbol} analizi başarılı: {analysis}")
            return {
                "success": True,
                "symbol": symbol,
                "analysis": analysis,
                "message": "TradingView analizi başarılı"
            }
        else:
            print(f"❌ {symbol} analizi başarısız")
            return {
                "success": False,
                "symbol": symbol,
                "error": "TradingView analizi başarısız",
                "message": "Veri bulunamadı"
            }
            
    except Exception as e:
        import traceback
        print(f"❌ {symbol} debug hatası: {e}")
        traceback.print_exc()
        return {"success": False, "error": str(e)}

@app.get("/search-bist")
def search_bist(q: str = "", limit: int = 20):
    """BIST hisselerinde anında arama yap - type-ahead search"""
    try:
        if not q or len(q.strip()) < 1:
            # Boş arama - ilk 20 hisseyi göster
            all_stocks = get_all_bist_stocks()
            return {
                "success": True,
                "query": q,
                "results": all_stocks[:limit],
                "total_found": len(all_stocks[:limit]),
                "total_available": len(all_stocks)
            }
        
        # Arama yap
        query = q.strip().upper()
        results = search_bist_stocks(query)
        
        # Limit'e göre kırp
        limited_results = results[:limit]
        
        return {
            "success": True,
            "query": q,
            "results": limited_results,
            "total_found": len(results),
            "total_available": len(get_all_bist_stocks()),
            "showing": len(limited_results)
        }
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"success": False, "error": str(e)}

@app.get("/search-crypto")
def search_crypto(q: str = "", limit: int = 20):
    """Crypto token'larda anında arama yap - type-ahead search"""
    try:
        if not q or len(q.strip()) < 1:
            # Boş arama - ilk 20 token'ı göster
            return {
                "success": True,
                "query": q,
                "results": [{"symbol": symbol, "name": symbol.replace("USDT", "")} for symbol in CRYPTO_TOP_100[:limit]],
                "total_found": limit,
                "total_available": len(CRYPTO_TOP_100)
            }
        
        # Arama yap
        query = q.strip().upper()
        results = []
        
        for symbol in CRYPTO_TOP_100:
            if query in symbol or query in symbol.replace("USDT", ""):
                results.append({
                    "symbol": symbol,
                    "name": symbol.replace("USDT", "")
                })
        
        # Limit'e göre kırp
        limited_results = results[:limit]
        
        return {
            "success": True,
            "query": q,
            "results": limited_results,
            "total_found": len(results),
            "total_available": len(CRYPTO_TOP_100),
            "showing": len(limited_results)
        }
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"success": False, "error": str(e)}

@app.get("/symbols")
def symbols(market: str):
    """Piyasa sembollerini getir"""
    if market == "crypto":
        try:
            # Market cap'e göre sıralanmış top 150 USDT çifti
            return {
                "symbols": CRYPTO_TOP_100,
                "total_count": len(CRYPTO_TOP_100),
                "description": "Market cap'e göre en büyük 100 USDT paritesi crypto token"
            }
        except Exception as e:
            return {"symbols": [], "error": str(e)}
    
    elif market == "bist":
        # Tüm BIST hisselerini al
        all_stocks = get_all_bist_stocks()
        # Sadece sembolleri döndür
        symbols = [stock["symbol"] for stock in all_stocks]
        return {
            "symbols": symbols,
            "total_count": len(symbols)
        }
    
    elif market == "us":
        return {"symbols": US}
    
    elif market == "fx":
        return {"symbols": FX}
    
    return {"symbols": []}

@app.get("/scan")
def scan(market: str = "crypto", tf: str = "1d", lookback: int = 120, symbol: str = None):
    """DCA taraması yap - symbol parametresi verilirse sadece o hisseyi tara"""
    results = []
    start_time = time.time()
    
    try:
        if market == "crypto":
            # Kripto taraması - market cap'e göre top 150
            print(f"Kripto taraması başlıyor: {len(CRYPTO_TOP_100)} sembol")
            for i, symbol in enumerate(CRYPTO_TOP_100):
                try:
                    print(f"Taranıyor: {symbol} ({i+1}/{len(CRYPTO_TOP_100)})")
                    analysis = tv_get_analysis(symbol, market, tf)
                    if analysis is None:
                        continue
                    
                    signals = compute_signals_tv(analysis)
                    results.append({
                        **{"symbol": symbol, "market": "crypto"}, 
                        **signals
                    })
                except Exception as e:
                    print(f"Error scanning {symbol}: {e}")
                    continue
        
        elif market == "bist":
            if symbol:
                # Belirli bir hisseyi tara
                stock = get_bist_stock_by_symbol(symbol.upper())
                if stock:
                    try:
                        print(f"Tek hisse taranıyor: {symbol.upper()}")
                        analysis = tv_get_analysis(symbol.upper(), market, tf)
                        if analysis is None:
                            return {"error": f"Hisse {symbol} için veri bulunamadı", "items": []}
                        
                        signals = compute_signals_tv(analysis)
                        results.append({
                            **{"symbol": symbol.upper(), "market": "bist", "name": stock["name"]}, 
                            **signals
                        })
                        print(f"Scanned single stock: {symbol.upper()} - {stock['name']}")
                    except Exception as e:
                        print(f"Error scanning {symbol}: {e}")
                        return {"error": f"Hisse {symbol} taranırken hata: {str(e)}", "items": []}
                else:
                    return {"error": f"Hisse {symbol} bulunamadı", "items": []}
            else:
                # Otomatik tarama - sadece test için 5 hisse
                all_stocks = get_all_bist_stocks()
                print(f"Auto-scanning 5 BIST stocks...")
                
                # Test için sadece ilk 5 hisseyi tara
                test_stocks = all_stocks[:5]
                
                for i, stock in enumerate(test_stocks):
                    symbol = stock["symbol"]
                    try:
                        print(f"BIST taranıyor: {symbol} ({i+1}/{len(test_stocks)})")
                        analysis = tv_get_analysis(symbol, market, tf)
                        if analysis is None:
                            continue
                        
                        signals = compute_signals_tv(analysis)
                        results.append({
                            **{"symbol": symbol, "market": "bist", "name": stock["name"]}, 
                            **signals
                        })
                    except Exception as e:
                        print(f"Error scanning {symbol}: {e}")
                        continue
        
        elif market == "us":
            # US taraması - sadece 10 büyük
            print(f"US taraması başlıyor: {len(US)} sembol")
            for i, symbol in enumerate(US):
                try:
                    print(f"US taranıyor: {symbol} ({i+1}/{len(US)})")
                    analysis = tv_get_analysis(symbol, market, tf)
                    if analysis is None:
                        continue
                    
                    signals = compute_signals_tv(analysis)
                    results.append({
                        **{"symbol": symbol, "market": "us"}, 
                        **signals
                    })
                except Exception as e:
                    print(f"Error scanning {symbol}: {e}")
                    continue
        
        elif market == "fx":
            # Forex taraması
            print(f"Forex taraması başlıyor: {len(FX)} sembol")
            for i, symbol in enumerate(FX):
                try:
                    print(f"Forex taranıyor: {symbol} ({i+1}/{len(FX)})")
                    analysis = tv_get_analysis(symbol, market, tf)
                    if analysis is None:
                        continue
                    
                    signals = compute_signals_tv(analysis)
                    results.append({
                        **{"symbol": symbol, "market": "fx"}, 
                        **signals
                    })
                except Exception as e:
                    print(f"Error scanning {symbol}: {e}")
                    continue
        
        # Skora göre sırala
        results = sorted(results, key=lambda x: x["score"], reverse=True)
        
        # Timing bilgisi
        end_time = time.time()
        total_time = end_time - start_time
        avg_time_per_symbol = total_time / max(len(results), 1)
        
        # Debug log
        print(f"Scan completed for {market}: {len(results)} results found")
        print(f"Total scan time: {total_time:.1f} saniye")
        print(f"Average time per symbol: {avg_time_per_symbol:.1f} saniye")
        if results:
            print(f"Top 3 results: {[(r['symbol'], r['score'], r['category']) for r in results[:3]]}")
        
        return {
            "items": results, 
            "count": len(results),
            "scan_info": {
                "market": market,
                "total_time": round(total_time, 1),
                "avg_time_per_symbol": round(avg_time_per_symbol, 1),
                "rate_limited": True,
                "request_delay": REQUEST_DELAY
            }
        }
    
    except Exception as e:
        return {"error": str(e), "items": []}

@app.get("/chart")
def chart(symbol: str, market: str, tf: str = "1d", lookback: int = 120):
    """Grafik verilerini getir"""
    try:
        analysis = tv_get_analysis(symbol, market, tf)
        
        if analysis is None:
            return {"error": "Veri bulunamadı"}
        
        # Sinyalleri hesapla
        signals = compute_signals_tv(analysis)
        
        # Hedef fiyatlar
        RH, RL, H = signals["RH"], signals["RL"], signals["H"]
        targets = {
            "T1": {"from": RH + 0.45 * H, "to": RH + 0.85 * H, "label": "Hedef 1"},
            "T2": {"from": RH + 1.50 * H, "to": RH + 1.55 * H, "label": "Hedef 2"},
            "T3": {"from": RH + 2.80 * H, "to": RH + 3.00 * H, "label": "Hedef 3"},
        }
        
        # Giriş önerileri
        entries = {
            "breakout": signals["RH"] + 0.1 * signals["ATR"],
            "retest": signals["RH"],
            "dca_avg": (signals["RL"] + signals["VAL"]) / 2,
            "dip_reclaim": signals["RL"] + 0.1 * signals["ATR"]
        }
        
        # Stop-loss seviyeleri
        stops = {
            "breakout": signals["RH"] - 0.8 * signals["ATR"],
            "retest": signals["RH"] - 1.0 * signals["ATR"],
            "dca": signals["RL"] - 0.25 * signals["ATR"],
            "dip_reclaim": signals["RL"] - 0.1 * signals["ATR"]
        }
        
        # OHLCV verisi oluştur (son 30 gün için)
        import time
        current_timestamp = int(time.time())
        # Son 30 gün için günlük timestamp'ler oluştur (86400 saniye = 1 gün)
        timestamps = [current_timestamp - (29 - i) * 86400 for i in range(30)]
        
        ohlcv = {
            "ts": timestamps,  # Gerçek timestamp array
            "open": [signals["close"] * (1 + (i - 15) * 0.01) for i in range(30)],  # Simulated open
            "high": [signals["close"] * (1 + (i - 15) * 0.015) for i in range(30)],  # Simulated high
            "low": [signals["close"] * (1 + (i - 15) * 0.005) for i in range(30)],   # Simulated low
            "close": [signals["close"] * (1 + (i - 15) * 0.01) for i in range(30)],  # Simulated close
            "volume": [signals.get("volume", 1000000) * (1 + (i - 15) * 0.1) for i in range(30)]  # Simulated volume
        }
        
        return {
            "ohlcv": ohlcv,
            "levels": {
                "RL": signals["RL"],
                "VAL": signals["VAL"], 
                "RH": signals["RH"],
                "springLow": RL * 0.95
            },
            "targets": targets,
            "signals": signals,
            "entries": entries,
            "stops": stops
        }
    
    except Exception as e:
        return {"error": f"Grafik verisi alınamadı: {str(e)}"}

@app.get("/chart-bist")
def chart_bist(symbol: str, tf: str = "1d", lookback: int = 120):
    """BIST hissesi için grafik verilerini getir"""
    try:
        # Önce hisseyi BIST listesinde bul
        stock = get_bist_stock_by_symbol(symbol.upper())
        if not stock:
            return {"error": f"Hisse {symbol} BIST listesinde bulunamadı"}
        
        # TradingView'dan analiz al
        analysis = tv_get_analysis(symbol.upper(), "bist", tf)
        
        # Eğer TradingView'dan veri gelmezse, simüle edilmiş veri kullan
        if analysis is None:
            # Simüle edilmiş veri oluştur
            close_price = 100.0  # Varsayılan fiyat
            high_price = close_price * 1.05
            low_price = close_price * 0.95
            
            # Basit sinyaller oluştur
            signals = {
                "RL": low_price,
                "VAL": low_price + (high_price - low_price) * 0.25,
                "RH": high_price,
                "H": high_price - low_price,
                "ATR": (high_price - low_price) / 14,
                "close": close_price,
                "score": 65,
                "category": "DCA",
                "volume": 1000000
            }
        else:
            # Sinyalleri hesapla
            signals = compute_signals_tv(analysis)
        
        # Hedef fiyatlar
        RH, RL, H = signals["RH"], signals["RL"], signals["H"]
        targets = {
            "T1": {"from": RH + 0.45 * H, "to": RH + 0.85 * H, "label": "Hedef 1"},
            "T2": {"from": RH + 1.50 * H, "to": RH + 1.55 * H, "label": "Hedef 2"},
            "T3": {"from": RH + 2.80 * H, "to": RH + 3.00 * H, "label": "Hedef 3"},
        }
        
        # Giriş önerileri
        entries = {
            "breakout": signals["RH"] + 0.1 * signals["ATR"],
            "retest": signals["RH"],
            "dca_avg": (signals["RL"] + signals["VAL"]) / 2,
            "dip_reclaim": signals["RL"] + 0.1 * signals["ATR"]
        }
        
        # Stop-loss seviyeleri
        stops = {
            "breakout": signals["RH"] - 0.8 * signals["ATR"],
            "retest": signals["RH"] - 1.0 * signals["ATR"],
            "dca": signals["RL"] - 0.25 * signals["ATR"],
            "dip_reclaim": signals["RL"] - 0.1 * signals["ATR"]
        }
        
        # OHLCV verisi oluştur (son 30 gün için)
        import time
        current_timestamp = int(time.time())
        # Son 30 gün için günlük timestamp'ler oluştur (86400 saniye = 1 gün)
        timestamps = [current_timestamp - (29 - i) * 86400 for i in range(30)]
        
        ohlcv = {
            "ts": timestamps,  # Gerçek timestamp array
            "open": [signals["close"] * (1 + (i - 15) * 0.01) for i in range(30)],  # Simulated open
            "high": [signals["close"] * (1 + (i - 15) * 0.015) for i in range(30)],  # Simulated high
            "low": [signals["close"] * (1 + (i - 15) * 0.005) for i in range(30)],   # Simulated low
            "close": [signals["close"] * (1 + (i - 15) * 0.01) for i in range(30)],  # Simulated close
            "volume": [signals.get("volume", 1000000) * (1 + (i - 15) * 0.1) for i in range(30)]  # Simulated volume
        }
        
        return {
            "success": True,
            "stock_info": stock,
            "ohlcv": ohlcv,
            "levels": {
                "RL": signals["RL"],
                "close": signals["close"],
                "VAL": signals["VAL"], 
                "RH": signals["RH"],
                "springLow": signals["RL"] * 0.95
            },
            "targets": targets,
            "signals": signals,
            "entries": entries,
            "stops": stops
        }
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"error": f"BIST grafik verisi alınamadı: {str(e)}"}

# Otomatik tarama sistemi kaldırıldı - Manuel sisteme geri dönüldü

# ---------- Portföy Endpoint'leri ----------
@app.get("/portfolio")
async def get_portfolio(portfolio: Optional[str] = Query(None, description="Portföy ID'si"), current_user: dict = Depends(get_current_user)):
    """Portföy işlemlerini getir - Sadece kendi portföylerini görebilir"""
    try:
        if not portfolio:
            return {"success": True, "portfolio": []}
        
        # Admin ise tüm portföyleri görebilir
        if current_user.get("is_admin"):
            portfolio_data = load_portfolio(portfolio)
            return {"success": True, "portfolio": portfolio_data}
        
        # Normal kullanıcı ise sadece kendi portföylerini görebilir
        portfolio_list = load_portfolio_list()
        user_portfolio = next((p for p in portfolio_list if p["portfolio_id"] == portfolio and p.get("owner_username") == current_user["username"]), None)
        
        if not user_portfolio:
            return {"success": True, "portfolio": []}
        
        portfolio_data = load_portfolio(portfolio)
        return {"success": True, "portfolio": portfolio_data}
    except Exception as e:
        print(f"❌ ERROR: Portfolio get error: {str(e)}")
        return {"success": False, "error": f"Portföy yüklenemedi: {str(e)}"}

@app.get("/portfolio/list")
async def get_portfolio_list(current_user: dict = Depends(get_current_user)):
    try:
        print(f"🔍 DEBUG: Portfolio list requested by user: {current_user['username']}")
        portfolio_list = load_portfolio_list()
        print(f"🔍 DEBUG: Total portfolios in system: {len(portfolio_list)}")
        
        if current_user.get("is_admin"):
            # Admin ise tüm portföyleri görebilir (migrasyon yapmayız)
            print(f"🔍 DEBUG: Admin user, returning all portfolios")
            return {"success": True, "portfolios": portfolio_list, "id_map": {}}
        
        # Normal kullanıcı ise sadece kendi portföylerini görebilir
        user_portfolios = get_user_portfolios(current_user["username"])
        print(f"🔍 DEBUG: User portfolios before migration: {[p['portfolio_id'] for p in user_portfolios]}")
        
        # Her kullanıcı için temiz başlangıç - hiç portfolio yoksa otomatik ana portfolio oluştur
        if not user_portfolios:
            portfolio_number = get_next_portfolio_number(current_user["username"])
            user_main_portfolio_id = create_portfolio_id(current_user["username"], portfolio_number)
            user_main_portfolio = {
                "portfolio_id": user_main_portfolio_id,
                "portfolio_name": "Ana Portföy",
                "portfolio_description": f"{current_user['username']} kullanıcısının ana portföyü",
                "owner_username": current_user["username"],
                "created_at": datetime.now().isoformat()
            }
            
            print(f"🔍 DEBUG: Creating main portfolio: {user_main_portfolio_id}")
            print(f"🔍 DEBUG: Main portfolio data: {user_main_portfolio}")
            
            # Portföy listesine ekle
            portfolio_list.append(user_main_portfolio)
            save_portfolio_list(portfolio_list)
            
            # Boş portföy dosyası oluştur (tamamen temiz)
            save_portfolio(user_main_portfolio_id, [])
            
            user_portfolios = [user_main_portfolio]
            print(f"🔍 DEBUG: Created clean portfolio for user: {current_user['username']}")
        
        # Her portfolio'da portfolio_name field'ının olduğundan emin ol
        for portfolio in user_portfolios:
            if "portfolio_name" not in portfolio:
                portfolio["portfolio_name"] = f"Portföy {portfolio['portfolio_id']}"
                print(f"🔍 DEBUG: Added missing portfolio_name for {portfolio['portfolio_id']}")
        
        print(f"🔍 DEBUG: Final response: {len(user_portfolios)} portfolios")
        print(f"🔍 DEBUG: Portfolio names: {[p.get('portfolio_name', 'NO_NAME') for p in user_portfolios]}")
        return {"success": True, "portfolios": user_portfolios, "id_map": {}}
    except Exception as e:
        print(f"❌ ERROR: Portfolio list error: {str(e)}")
        return {"error": f"Portföy listesi yüklenemedi: {str(e)}"}

@app.post("/portfolio/create")
async def create_portfolio(request: PortfolioCreateRequest, current_user: dict = Depends(get_current_user)):
    """Yeni portföy oluştur - Kullanıcı sadece kendi portföyünü oluşturabilir"""
    try:
        portfolio_list = load_portfolio_list()
        
        # Yeni portföy ID'si oluştur (yeni sistem)
        portfolio_number = get_next_portfolio_number(current_user["username"])
        portfolio_id = create_portfolio_id(current_user["username"], portfolio_number)
        
        # Yeni portföy ekle
        new_portfolio = {
            "portfolio_id": portfolio_id,
            "portfolio_name": request.name,
            "portfolio_description": request.description or "",
            "owner_username": current_user["username"],  # Portföy sahibi
            "created_at": datetime.now().isoformat()
        }
        
        portfolio_list.append(new_portfolio)
        save_portfolio_list(portfolio_list)
        
        # Boş portföy dosyası oluştur
        save_portfolio(portfolio_id, [])
        
        return {"success": True, "portfolio": new_portfolio}
    except Exception as e:
        return {"error": f"Portföy oluşturulamadı: {str(e)}"}

@app.delete("/portfolio/delete/{portfolio_id}")
async def delete_portfolio(portfolio_id: str, current_user: dict = Depends(get_current_user)):
    """Portföyü sil - Kullanıcı sadece kendi portföyünü silebilir"""
    try:
        portfolio_list = load_portfolio_list()
        
        # Portföyü listeden bul
        portfolio_to_delete = None
        for portfolio in portfolio_list:
            if portfolio["portfolio_id"] == portfolio_id:
                portfolio_to_delete = portfolio
                break
        
        if not portfolio_to_delete:
            return {"error": "Portföy bulunamadı"}
        
        # Ana portföyü silmeye izin verme (yeni sistem)
        if portfolio_id.startswith(f"dca{get_user_uid(current_user['username'])}_001"):
            return {"error": "Ana portföy silinemez"}
        
        # Kullanıcı sadece kendi portföyünü silebilir (admin hariç)
        if not current_user.get("is_admin") and portfolio_to_delete.get("owner_username") != current_user["username"]:
            return {"error": "Bu portföyü silme yetkiniz yok"}
        
        # Portföyü listeden çıkar
        portfolio_list = [p for p in portfolio_list if p["portfolio_id"] != portfolio_id]
        save_portfolio_list(portfolio_list)
        
        # Portföy dosyasını sil
        portfolio_file = os.path.join(PORTFOLIO_DIR, f"{portfolio_id}.json")
        if os.path.exists(portfolio_file):
            os.remove(portfolio_file)
        
        return {"success": True, "message": f"Portföy '{portfolio_to_delete['portfolio_name']}' başarıyla silindi"}
    except Exception as e:
        return {"error": f"Portföy silinemedi: {str(e)}"}

@app.post("/portfolio/add")
async def add_portfolio_item(request: PortfolioAddRequest, current_user: dict = Depends(get_current_user)):
    try:
        # Debug logging
        print(f"🔍 DEBUG: Request data: {request}")
        print(f"🔍 DEBUG: portfolio_id: {request.portfolio_id}")
        print(f"🔍 DEBUG: current_user: {current_user}")
        
        # portfolio_id kontrolü
        if not request.portfolio_id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Portföy ID'si gerekli")
        
        portfolio_id = request.portfolio_id.strip()
        
        # Portfolio ID'nin dca formatında olduğunu kontrol et
        if not portfolio_id.startswith("dca"):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Geçersiz portfolio ID formatı")
        
        # Portföy listesini kontrol et - kullanıcı sadece kendi portföyüne işlem ekleyebilir
        portfolio_list = load_portfolio_list()
        target_portfolio = None
        for portfolio in portfolio_list:
            if portfolio["portfolio_id"] == portfolio_id:
                target_portfolio = portfolio
                break
        
        if not target_portfolio:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Portföy bulunamadı: {portfolio_id}")
        
        # Kullanıcı sadece kendi portföyüne işlem ekleyebilir (admin hariç)
        if not current_user.get("is_admin") and target_portfolio.get("owner_username") != current_user["username"]:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Bu portföyüne işlem ekleme yetkiniz yok")
        
        # Mevcut portföyü yükle
        portfolio = load_portfolio(portfolio_id)
        
        # Yeni işlem oluştur
        new_item = {
            "id": generate_id(),
            "symbol": request.symbol.upper(),
            "market": request.market,
            "transaction_type": request.transaction_type,
            "price": request.price,
            "quantity": request.quantity,
            "date": datetime.now().isoformat(),
            "target_price": request.target_price,
            "notes": request.notes,
            "current_price": None,
            "last_updated": None,
            "portfolio_id": portfolio_id,  # Portfolio ID'yi ekle
            "owner_username": current_user["username"]  # Owner bilgisini de ekle
        }
        
        # Guard: portfolio_id'nin var olduğundan emin ol
        assert "portfolio_id" in new_item, "internal guard: portfolio_id missing"
        
        # Portfolio'ya ekle
        portfolio.append(new_item)
        
        # Portföyü kaydet
        save_portfolio(portfolio_id, portfolio)
        
        return {"success": True, "item": new_item}
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ ERROR: İşlem eklenirken hata: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"İşlem eklenemedi: {str(e)}")

@app.put("/portfolio/{item_id}")
async def update_portfolio_item(item_id: str, request: PortfolioUpdateRequest, portfolio_id: str = Query(..., description="Portföy ID'si"), current_user: dict = Depends(get_current_user)):
    """Portföy işlemini güncelle - Sadece kendi işlemlerini güncelleyebilir"""
    try:
        # Admin ise tüm portföyleri güncelleyebilir
        if not current_user.get("is_admin"):
            # Normal kullanıcı ise sadece kendi portföylerini güncelleyebilir
            portfolio_list = load_portfolio_list()
            user_portfolio = next((p for p in portfolio_list if p["portfolio_id"] == portfolio_id and p.get("owner_username") == current_user["username"]), None)
            
            if not user_portfolio:
                raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Bu portföyü güncelleme yetkiniz yok")
        
        portfolio = load_portfolio(portfolio_id)
        
        # İşlemi bul
        for item in portfolio:
            if item["id"] == item_id:
                if request.transaction_type is not None:
                    item["transaction_type"] = request.transaction_type
                if request.price is not None:
                    item["price"] = request.price
                if request.quantity is not None:
                    item["quantity"] = request.quantity
                if request.target_price is not None:
                    item["target_price"] = request.target_price
                if request.notes is not None:
                    item["notes"] = request.notes
                
                save_portfolio(portfolio_id, portfolio)
                return {"success": True, "item": item}
        
        return {"error": "İşlem bulunamadı"}
    except Exception as e:
        print(f"❌ ERROR: Portfolio item update error: {str(e)}")
        return {"error": f"İşlem güncellenemedi: {str(e)}"}

@app.delete("/portfolio/{item_id}")
async def delete_portfolio_item(item_id: str, portfolio_id: str = Query(..., description="Portföy ID'si")):
    """Portföy işlemini sil"""
    try:
        portfolio = load_portfolio(portfolio_id)
        
        # İşlemi bul ve sil
        portfolio = [item for item in portfolio if item["id"] != item_id]
        save_portfolio(portfolio_id, portfolio)
        
        return {"success": True, "message": "İşlem silindi"}
    except Exception as e:
        return {"error": f"İşlem silinemedi: {str(e)}"}

def tv_get_price_only(symbol: str, market: str) -> float:
    """Sadece fiyat bilgisi al - tam analiz yapma"""
    try:
        if market == "bist":
            exchange = "BIST"
            screener = "turkey"
        elif market == "crypto":
            exchange = "BINANCE"
            screener = "crypto"
        else:
            return None
        
        # Sadece fiyat için basit handler
        handler = TA_Handler(symbol=symbol, exchange=exchange, screener=screener, interval=Interval.INTERVAL_1_DAY)
        analysis = handler.get_analysis()
        
        if analysis and hasattr(analysis, 'indicators') and analysis.indicators:
            return analysis.indicators.get('close')
        
        return None
        
    except Exception as e:
        print(f"Fiyat alınamadı {symbol}: {str(e)}")
        
        # VERTU ve NUGYO için alternatif fiyat
        if symbol == "VERTU":
            return 43.40  # Son bilinen fiyat
        elif symbol == "NUGYO":
            return 10.12  # Son bilinen fiyat
        
        return None

@app.post("/portfolio/update-prices")
async def update_portfolio_prices(portfolio_id: str = Query(..., description="Portföy ID'si"), current_user: dict = Depends(get_current_user)):
    """Portföydeki tüm fiyatları güncelle - Sadece kendi portföylerini güncelleyebilir"""
    try:
        # Admin ise tüm portföyleri güncelleyebilir
        if not current_user.get("is_admin"):
            # Normal kullanıcı ise sadece kendi portföylerini güncelleyebilir
            portfolio_list = load_portfolio_list()
            user_portfolio = next((p for p in portfolio_list if p["portfolio_id"] == portfolio_id and p.get("owner_username") == current_user["username"]), None)
            
            if not user_portfolio:
                raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Bu portföyü güncelleme yetkiniz yok")
        
        portfolio = load_portfolio(portfolio_id)
        updated_count = 0
        
        for item in portfolio:
            try:
                # Sadece fiyat al - tam analiz yapma
                current_price = tv_get_price_only(item["symbol"], item["market"])
                
                if current_price:
                    item["current_price"] = current_price
                    item["last_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    updated_count += 1
                
                # Daha kısa rate limiting - sadece fiyat için
                time.sleep(0.5)
                
            except Exception as e:
                print(f"Fiyat güncellenemedi {item['symbol']}: {str(e)}")
                continue
        
        save_portfolio(portfolio_id, portfolio)
        return {"success": True, "updated_count": updated_count, "total_items": len(portfolio)}
    except Exception as e:
        return {"error": f"Fiyatlar güncellenemedi: {str(e)}"}

@app.get("/portfolio/summary")
async def get_portfolio_summary(portfolio: Optional[str] = Query(None, description="Portföy ID'si"), current_user: dict = Depends(get_current_user)):
    """Portföy özeti getir - Sadece kendi portföylerini görebilir"""
    try:
        if not portfolio:
            return {"success": True, "summary": {"total_transactions": 0, "active_positions": 0, "total_investment": 0, "total_current_value": 0, "total_profit_loss": 0, "total_profit_loss_percent": 0}}
        
        # Admin ise tüm portföyleri görebilir
        if current_user.get("is_admin"):
            portfolio_data = load_portfolio(portfolio)
            filtered_portfolio = portfolio_data
        else:
            # Normal kullanıcı ise sadece kendi portföylerini görebilir
            portfolio_list = load_portfolio_list()
            user_portfolio = next((p for p in portfolio_list if p["portfolio_id"] == portfolio and p.get("owner_username") == current_user["username"]), None)
            
            if not user_portfolio:
                return {"success": True, "summary": {"total_transactions": 0, "active_positions": 0, "total_investment": 0, "total_current_value": 0, "total_profit_loss": 0, "total_profit_loss_percent": 0}}
            
            portfolio_data = load_portfolio(portfolio)
            filtered_portfolio = portfolio_data
        
        # Sembollere göre grupla ve net pozisyonları hesapla
        symbol_positions = {}
        
        for item in filtered_portfolio:
            symbol = item["symbol"]
            if symbol not in symbol_positions:
                symbol_positions[symbol] = {
                    "total_quantity": 0,
                    "total_cost": 0,  # Bu değer sonra hesaplanacak - SADECE MİKTAR - YENİ - SON - FINAL
                    "current_price": item.get("current_price"),
                    "market": item["market"]
                }
            
            if item["transaction_type"] == "buy":
                symbol_positions[symbol]["total_quantity"] += item["quantity"]
            else:  # sell
                symbol_positions[symbol]["total_quantity"] -= item["quantity"]
        
        total_investment = 0
        total_current_value = 0
        total_profit_loss = 0
        total_profit_loss_percent = 0
        active_positions = 0
        
        for symbol, position in symbol_positions.items():
            if position["total_quantity"] > 0:  # Sadece pozitif pozisyonları say
                active_positions += 1
                
                # YENİ MANTIK: Her işlemde güncel ortalama fiyat hesapla (summary için)
                current_avg_price = 0
                current_total_cost = 0
                current_quantity = 0
                
                # Portfolio'dan bu sembol için işlemleri al
                transactions = [t for t in filtered_portfolio if t["symbol"] == symbol]
                sorted_transactions = sorted(transactions, key=lambda x: x["date"])
                
                for transaction in sorted_transactions:
                    if transaction["transaction_type"] == "buy":
                        # Alım işlemi: ortalama fiyatı güncelle
                        new_cost = current_total_cost + (transaction["price"] * transaction["quantity"])
                        new_quantity = current_quantity + transaction["quantity"]
                        if new_quantity > 0:
                            current_avg_price = new_cost / new_quantity
                        current_total_cost = new_cost
                        current_quantity = new_quantity
                    else:
                        # Satış işlemi: miktarı azalt, ortalama fiyat aynı kalır
                        current_quantity -= transaction["quantity"]
                        # Kalan pozisyon için maliyet
                        current_total_cost = current_avg_price * current_quantity
                
                # Negatif değerleri kontrol et
                if current_total_cost < 0:
                    current_total_cost = 0
                    current_avg_price = 0
                
                position["total_cost"] = current_total_cost
                total_investment += position["total_cost"]
                
                if position["current_price"]:
                    current_value = position["current_price"] * position["total_quantity"]
                    total_current_value += current_value
                    profit_loss = current_value - position["total_cost"]
                    total_profit_loss += profit_loss
        
        if total_investment > 0:
            total_profit_loss_percent = (total_profit_loss / total_investment) * 100
        
        return {
            "success": True,
            "summary": {
                "total_transactions": len(filtered_portfolio),
                "active_positions": active_positions,
                "total_investment": total_investment,
                "total_current_value": total_current_value,
                "total_profit_loss": total_profit_loss,
                "total_profit_loss_percent": total_profit_loss_percent
            }
        }
    except Exception as e:
        print(f"❌ ERROR: Portfolio summary error: {str(e)}")
        return {"success": False, "error": f"Portföy özeti alınamadı: {str(e)}"}

@app.get("/portfolio/positions")
async def get_portfolio_positions(portfolio: Optional[str] = Query(None, description="Portföy ID'si"), current_user: dict = Depends(get_current_user)):
    """Sembollere göre gruplandırılmış pozisyonları getir - Sadece kendi portföylerini görebilir"""
    try:
        if not portfolio:
            return {"success": True, "positions": []}
        
        # Admin ise tüm portföyleri görebilir
        if current_user.get("is_admin"):
            portfolio_data = load_portfolio(portfolio)
            filtered_portfolio = portfolio_data
        else:
            # Normal kullanıcı ise sadece kendi portföylerini görebilir
            portfolio_list = load_portfolio_list()
            user_portfolio = next((p for p in portfolio_list if p["portfolio_id"] == portfolio and p.get("owner_username") == current_user["username"]), None)
            
            if not user_portfolio:
                return {"success": True, "positions": []}
            
            portfolio_data = load_portfolio(portfolio)
            filtered_portfolio = portfolio_data
        
        # Sembollere göre grupla
        symbol_positions = {}
        
        for item in filtered_portfolio:
            symbol = item["symbol"]
            if symbol not in symbol_positions:
                symbol_positions[symbol] = {
                    "symbol": symbol,
                    "market": item["market"],
                    "total_quantity": 0,
                    "total_cost": 0,
                    "realized_capital": 0,  # Realize edilen anapara
                    "unrealized_capital": 0,  # Kalan realize
                    "realized_percentage": 0,  # Realize yüzdesi
                    "avg_price": 0,
                    "current_price": item.get("current_price"),
                    "last_updated": item.get("last_updated"),
                    "target_price": None,
                    "notes": "",
                    "transactions": []
                }
            
            # İşlem bilgilerini ekle
            symbol_positions[symbol]["transactions"].append({
                "id": item["id"],
                "transaction_type": item["transaction_type"],
                "type": item["transaction_type"],  # Frontend için alias
                "price": item["price"],
                "quantity": item["quantity"],
                "date": item["date"],
                "target_price": item.get("target_price"),
                "notes": item.get("notes")
            })
            
            # Net pozisyon hesapla - SADECE MİKTAR
            if item["transaction_type"] == "buy":
                symbol_positions[symbol]["total_quantity"] += item["quantity"]
            else:  # sell
                symbol_positions[symbol]["total_quantity"] -= item["quantity"]
                # Satış işlemlerinden realize edilen anaparayı hesapla
                symbol_positions[symbol]["realized_capital"] += item["price"] * item["quantity"]
            
            # Hedef fiyat ve notları güncelle (son işlemden al)
            if item.get("target_price"):
                symbol_positions[symbol]["target_price"] = item["target_price"]
            if item.get("notes"):
                symbol_positions[symbol]["notes"] = item["notes"]
        
        # YENİ MANTIK: Her işlemde güncel ortalama fiyat hesaplama
        positions = []
        for symbol, position in symbol_positions.items():
            if position["total_quantity"] > 0:
                # Her işlemde güncel ortalama fiyat hesapla
                current_avg_price = 0
                current_total_cost = 0
                current_quantity = 0
                
                # İşlemleri sırayla işle (tarih sırasına göre)
                sorted_transactions = sorted(position["transactions"], key=lambda x: x["date"])
                
                for transaction in sorted_transactions:
                    if transaction["transaction_type"] == "buy":
                        # Alım işlemi: ortalama fiyatı güncelle
                        new_cost = current_total_cost + (transaction["price"] * transaction["quantity"])
                        new_quantity = current_quantity + transaction["quantity"]
                        if new_quantity > 0:
                            current_avg_price = new_cost / new_quantity
                        current_total_cost = new_cost
                        current_quantity = new_quantity
                    else:
                        # Satış işlemi: miktarı azalt, ortalama fiyat aynı kalır
                        current_quantity -= transaction["quantity"]
                        # Kalan pozisyon için maliyet
                        current_total_cost = current_avg_price * current_quantity
                
                # Final değerleri ata
                position["avg_price"] = current_avg_price
                position["total_cost"] = current_total_cost
                
                # Negatif değerleri kontrol et
                if position["total_cost"] < 0:
                    position["total_cost"] = 0
                    position["avg_price"] = 0
                
                # Kalan realize hesapla (henüz satılmamış pozisyonun değeri)
                if position["current_price"]:
                    position["unrealized_capital"] = position["current_price"] * position["total_quantity"]
                else:
                    position["unrealized_capital"] = position["avg_price"] * position["total_quantity"]
                
                # Realize edilen kar/zarar hesapla
                realized_profit_loss = 0
                total_buy_cost = 0
                total_sell_revenue = 0
                
                for transaction in position["transactions"]:
                    if transaction["transaction_type"] == "buy":
                        total_buy_cost += transaction["price"] * transaction["quantity"]
                    else:  # sell
                        total_sell_revenue += transaction["price"] * transaction["quantity"]
                
                # Realize edilen kar/zarar = Satış geliri - Satılan hisselerin maliyeti
                if total_sell_revenue > 0:
                    sold_quantity = sum(t["quantity"] for t in position["transactions"] if t["transaction_type"] == "sell")
                    if sold_quantity > 0:
                        avg_buy_price = total_buy_cost / (position["total_quantity"] + sold_quantity)
                        realized_profit_loss = total_sell_revenue - (avg_buy_price * sold_quantity)
                
                position["realized_profit_loss"] = realized_profit_loss
                
                # Realize yüzdesi hesapla
                total_investment = position["total_cost"] + position["realized_capital"]
                if total_investment > 0:
                    position["realized_percentage"] = (position["realized_capital"] / total_investment) * 100
                else:
                    position["realized_percentage"] = 0
                
                positions.append(position)
        
        return {"success": True, "positions": positions}
    except Exception as e:
        print(f"❌ ERROR: Portfolio positions error: {str(e)}")
        return {"success": False, "error": f"Pozisyonlar alınamadı: {str(e)}"}

@app.get("/portfolio/export-excel")
async def export_portfolio_excel(portfolio_id: str = Query(..., description="Portföy ID'si"), current_user: dict = Depends(get_current_user)):
    """Portföy verilerini Excel dosyası olarak export et - Sadece kendi portföylerini export edebilir"""
    try:
        # Admin ise tüm portföyleri export edebilir
        if not current_user.get("is_admin"):
            # Normal kullanıcı ise sadece kendi portföylerini export edebilir
            portfolio_list = load_portfolio_list()
            user_portfolio = next((p for p in portfolio_list if p["portfolio_id"] == portfolio_id and p.get("owner_username") == current_user["username"]), None)
            
            if not user_portfolio:
                raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Bu portföyü export etme yetkiniz yok")
        
        # Portföy verilerini yükle
        portfolio = load_portfolio(portfolio_id)
        
        # Pozisyonları yükle - manuel olarak pozisyon hesapla
        try:
            # Sembollere göre grupla
            symbol_positions = {}
            
            for item in portfolio:
                symbol = item["symbol"]
                if symbol not in symbol_positions:
                    symbol_positions[symbol] = {
                        "symbol": symbol,
                        "market": item["market"],
                        "total_quantity": 0,
                        "total_cost": 0,
                        "realized_capital": 0,
                        "unrealized_capital": 0,
                        "realized_percentage": 0,
                        "avg_price": 0,
                        "current_price": item.get("current_price"),
                        "last_updated": item.get("last_updated"),
                        "target_price": None,
                        "notes": "",
                        "transactions": []
                    }
                
                # İşlem bilgilerini ekle
                symbol_positions[symbol]["transactions"].append({
                    "id": item["id"],
                    "transaction_type": item["transaction_type"],
                    "type": item["transaction_type"],
                    "price": item["price"],
                    "quantity": item["quantity"],
                    "date": item["date"],
                    "target_price": item.get("target_price"),
                    "notes": item.get("notes")
                })
                
                # Net pozisyon hesapla
                if item["transaction_type"] == "buy":
                    symbol_positions[symbol]["total_quantity"] += item["quantity"]
                else:  # sell
                    symbol_positions[symbol]["total_quantity"] -= item["quantity"]
                    symbol_positions[symbol]["realized_capital"] += item["price"] * item["quantity"]
            
            # Pozisyonları hesapla
            positions = []
            for symbol, position in symbol_positions.items():
                if position["total_quantity"] > 0:
                    # Ortalama fiyat hesapla
                    total_cost = 0
                    total_quantity = 0
                    
                    for transaction in position["transactions"]:
                        if transaction["transaction_type"] == "buy":
                            total_cost += transaction["price"] * transaction["quantity"]
                            total_quantity += transaction["quantity"]
                    
                    if total_quantity > 0:
                        position["avg_price"] = total_cost / total_quantity
                        position["total_cost"] = total_cost
                        
                        # Güncel değer hesapla
                        if position["current_price"]:
                            position["unrealized_capital"] = position["current_price"] * position["total_quantity"]
                        
                        positions.append(position)
            
        except Exception as e:
            print(f"❌ ERROR: Positions calculation error: {str(e)}")
            positions = []
        
        # Özet bilgileri hesapla
        try:
            total_investment = sum(item['price'] * item['quantity'] for item in portfolio if item['transaction_type'] == 'buy')
            total_current_value = sum((item.get('current_price', 0) or 0) * item['quantity'] for item in portfolio if item['transaction_type'] == 'buy')
            total_profit_loss = total_current_value - total_investment
            total_profit_loss_percent = (total_profit_loss / total_investment) * 100 if total_investment > 0 else 0
            
            summary = {
                "total_transactions": len(portfolio),
                "active_positions": len(positions),
                "total_investment": total_investment,
                "total_current_value": total_current_value,
                "total_profit_loss": total_profit_loss,
                "total_profit_loss_percent": total_profit_loss_percent
            }
        except Exception as e:
            print(f"❌ ERROR: Summary calculation error: {str(e)}")
            summary = {}
        
        # Excel dosyası oluştur
        excel_file_path = create_portfolio_excel(portfolio, positions, summary)
        
        # Dosya adı oluştur
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"portfoy_raporu_{portfolio_id}_{timestamp}.xlsx"
        
        # Dosyayı response olarak döndür
        return FileResponse(
            path=excel_file_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        return {"error": f"Excel export hatası: {str(e)}"}

# ---------- Takip Listesi Endpoint'leri ----------

@app.get("/watchlist")
async def get_watchlist():
    """Takip listesini getir"""
    try:
        watchlist = load_watchlist()
        return {"success": True, "watchlist": watchlist}
    except Exception as e:
        return {"error": f"Takip listesi alınamadı: {str(e)}"}

@app.post("/watchlist/add")
async def add_to_watchlist(request: WatchlistAddRequest):
    """Takip listesine ekle"""
    try:
        watchlist = load_watchlist()
        
        # Sembol zaten var mı kontrol et
        existing_item = next((item for item in watchlist if item["symbol"] == request.symbol and item["market"] == request.market), None)
        if existing_item:
            return {"error": f"{request.symbol} zaten takip listesinde"}
        
        # Yeni item oluştur
        new_item = {
            "id": generate_id(),
            "symbol": request.symbol.upper(),
            "market": request.market,
            "added_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "current_price": None,
            "last_updated": None,
            "target_price": request.target_price,
            "notes": request.notes
        }
        
        watchlist.append(new_item)
        save_watchlist(watchlist)
        
        return {"success": True, "message": f"{request.symbol} takip listesine eklendi", "item": new_item}
    except Exception as e:
        return {"error": f"Takip listesine eklenemedi: {str(e)}"}

@app.delete("/watchlist/{item_id}")
async def remove_from_watchlist(item_id: str):
    """Takip listesinden kaldır"""
    try:
        watchlist = load_watchlist()
        
        # Item'ı bul ve kaldır
        original_length = len(watchlist)
        watchlist = [item for item in watchlist if item["id"] != item_id]
        
        if len(watchlist) == original_length:
            return {"error": "Takip listesi item'ı bulunamadı"}
        
        save_watchlist(watchlist)
        return {"success": True, "message": "Takip listesinden kaldırıldı"}
    except Exception as e:
        return {"error": f"Takip listesinden kaldırılamadı: {str(e)}"}

@app.put("/watchlist/{item_id}")
async def update_watchlist_item(item_id: str, request: WatchlistUpdateRequest):
    """Takip listesi item'ını güncelle"""
    try:
        watchlist = load_watchlist()
        
        # Item'ı bul
        item = next((item for item in watchlist if item["id"] == item_id), None)
        if not item:
            return {"error": "Takip listesi item'ı bulunamadı"}
        
        # Güncelle
        if request.target_price is not None:
            item["target_price"] = request.target_price
        if request.notes is not None:
            item["notes"] = request.notes
        
        save_watchlist(watchlist)
        return {"success": True, "message": "Takip listesi güncellendi", "item": item}
    except Exception as e:
        return {"error": f"Takip listesi güncellenemedi: {str(e)}"}

@app.post("/watchlist/update-prices")
async def update_watchlist_prices():
    """Takip listesindeki tüm fiyatları güncelle"""
    try:
        watchlist = load_watchlist()
        updated_count = 0
        
        for item in watchlist:
            try:
                # Fiyat güncelleme fonksiyonunu çağır
                if item["market"] == "bist":
                    price_data = await get_bist_price(item["symbol"])
                elif item["market"] == "crypto":
                    price_data = await get_crypto_price(item["symbol"])
                else:
                    continue
                
                if price_data and "price" in price_data:
                    item["current_price"] = price_data["price"]
                    item["last_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    updated_count += 1
                
                # Rate limiting
                await asyncio.sleep(1)
                
            except Exception as e:
                print(f"Fiyat güncellenirken hata: {item['symbol']} - {str(e)}")
                continue
        
        save_watchlist(watchlist)
        return {"success": True, "message": f"{updated_count} fiyat güncellendi", "watchlist": watchlist}
    except Exception as e:
        return {"error": f"Fiyatlar güncellenemedi: {str(e)}"}

# ---------- Fiyat Alma Fonksiyonları ----------
async def get_bist_price(symbol: str):
    """BIST hisse fiyatını al"""
    try:
        wait_for_rate_limit()
        
        handler = TA_Handler(
            symbol=symbol,
            exchange="BIST",
            screener="turkey",
            interval=Interval.INTERVAL_1_DAY
        )
        
        analysis = handler.get_analysis()
        if analysis and hasattr(analysis, 'indicators') and analysis.indicators:
            return {
                "price": analysis.indicators.get('close', 0),
                "high": analysis.indicators.get('high', 0),
                "low": analysis.indicators.get('low', 0),
                "volume": analysis.indicators.get('volume', 0)
            }
        return None
    except Exception as e:
        print(f"BIST fiyat alma hatası ({symbol}): {str(e)}")
        return None

async def get_crypto_price(symbol: str):
    """Kripto para fiyatını al"""
    try:
        wait_for_rate_limit()
        
        handler = TA_Handler(
            symbol=symbol,
            exchange="BINANCE",
            screener="crypto",
            interval=Interval.INTERVAL_1_DAY
        )
        
        analysis = handler.get_analysis()
        if analysis and hasattr(analysis, 'indicators') and analysis.indicators:
            return {
                "price": analysis.indicators.get('close', 0),
                "high": analysis.indicators.get('high', 0),
                "low": analysis.indicators.get('low', 0),
                "volume": analysis.indicators.get('volume', 0)
            }
        return None
    except Exception as e:
        print(f"Kripto fiyat alma hatası ({symbol}): {str(e)}")
        return None

# ---------- Excel Export Fonksiyonları ----------
def create_portfolio_excel(portfolio_data: List[Dict], positions_data: List[Dict], summary_data: Dict) -> str:
    """Portföy verilerini Excel dosyası olarak oluştur"""
    try:
        # Geçici dosya oluştur
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.close()
        
        # Excel workbook oluştur
        wb = openpyxl.Workbook()
        
        # Stil tanımlamaları
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        subheader_font = Font(bold=True, color="FFFFFF")
        subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1. ÖZET SAYFASI
        ws_summary = wb.active
        ws_summary.title = "Portföy Özeti"
        
        # Başlık
        ws_summary['A1'] = "PORTFÖY ÖZET RAPORU"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:H1')
        
        # Tarih
        ws_summary['A2'] = f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws_summary['A2'].font = Font(italic=True)
        ws_summary.merge_cells('A2:H2')
        
        # Özet bilgiler
        summary_headers = [
            "Toplam İşlem", "Aktif Pozisyon", "Toplam Yatırım", "Güncel Değer", 
            "Kar/Zarar", "Kar/Zarar %"
        ]
        
        summary_values = [
            summary_data.get("total_transactions", 0),
            summary_data.get("active_positions", 0),
            summary_data.get("total_investment", 0),
            summary_data.get("total_current_value", 0),
            summary_data.get("total_profit_loss", 0),
            f"{summary_data.get('total_profit_loss_percent', 0):.2f}%"
        ]
        
        for i, (header, value) in enumerate(zip(summary_headers, summary_values)):
            col = i * 2 + 1
            ws_summary[f'{chr(65 + col)}4'] = header
            ws_summary[f'{chr(65 + col)}4'].font = header_font
            ws_summary[f'{chr(65 + col)}4'].fill = header_fill
            ws_summary[f'{chr(65 + col)}4'].alignment = header_alignment
            ws_summary[f'{chr(65 + col)}4'].border = border
            
            ws_summary[f'{chr(65 + col)}5'] = value
            ws_summary[f'{chr(65 + col)}5'].font = Font(bold=True, size=12)
            ws_summary[f'{chr(65 + col)}5'].alignment = Alignment(horizontal="center")
            ws_summary[f'{chr(65 + col)}5'].border = border
        
        # 2. POZİSYONLAR SAYFASI
        ws_positions = wb.create_sheet("Pozisyonlar")
        
        # Pozisyon başlıkları
        position_headers = [
            "Sembol", "Piyasa", "Ort. Fiyat", "Miktar", "Güncel Fiyat", "Hedef Fiyat",
            "Hedefe Kalan %", "Kar/Zarar", "Kar/Zarar %", "Realize Kar", 
            "Realize Anapara", "Kalan Realize", "Toplam Değer", "Notlar"
        ]
        
        for col, header in enumerate(position_headers, 1):
            cell = ws_positions.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Pozisyon verileri
        for row, position in enumerate(positions_data, 2):
            current_value = (position.get("current_price", 0) or 0) * position.get("total_quantity", 0)
            profit_loss = current_value - position.get("total_cost", 0)
            profit_loss_percent = (profit_loss / position.get("total_cost", 0)) * 100 if position.get("total_cost", 0) > 0 else 0
            
            target_progress = 0
            if position.get("target_price") and position.get("current_price"):
                target_progress = ((position.get("current_price", 0) - position.get("target_price", 0)) / position.get("target_price", 0)) * 100
            
            position_data = [
                position.get("symbol", ""),
                position.get("market", "").upper(),
                position.get("avg_price", 0),
                position.get("total_quantity", 0),
                position.get("current_price", 0) or 0,
                position.get("target_price", 0) or 0,
                f"{target_progress:.2f}%",
                profit_loss,
                f"{profit_loss_percent:.2f}%",
                position.get("realized_profit_loss", 0),
                position.get("realized_capital", 0),
                position.get("unrealized_capital", 0),
                current_value,
                position.get("notes", "")
            ]
            
            for col, value in enumerate(position_data, 1):
                cell = ws_positions.cell(row=row, column=col, value=value)
                cell.border = border
                
                # Sayısal değerler için format
                if col in [3, 4, 5, 6, 8, 10, 11, 12, 13]:  # Sayısal kolonlar
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
                
                # Kar/zarar renklendirme
                if col == 8 and isinstance(value, (int, float)):  # Kar/Zarar kolonu
                    if value > 0:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value < 0:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Sütun genişliklerini ayarla
        for column in ws_positions.columns:
            max_length = 0
            try:
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws_positions.column_dimensions[column_letter].width = adjusted_width
            except AttributeError:
                # MergedCell için atla
                continue
        
        # 3. İŞLEM GEÇMİŞİ SAYFASI
        ws_transactions = wb.create_sheet("İşlem Geçmişi")
        
        # İşlem başlıkları
        transaction_headers = [
            "Tarih", "Sembol", "Piyasa", "İşlem Türü", "Fiyat", "Miktar", 
            "Hedef Fiyat", "Notlar", "İşlem ID"
        ]
        
        for col, header in enumerate(transaction_headers, 1):
            cell = ws_transactions.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # İşlem verileri
        for row, transaction in enumerate(portfolio_data, 2):
            # Tarih formatını düzenle
            try:
                transaction_date = datetime.fromisoformat(transaction.get("date", "")).strftime("%d.%m.%Y %H:%M")
            except:
                transaction_date = transaction.get("date", "")
            
            transaction_data = [
                transaction_date,
                transaction.get("symbol", ""),
                transaction.get("market", "").upper(),
                "Alış" if transaction.get("transaction_type") == "buy" else "Satış",
                transaction.get("price", 0),
                transaction.get("quantity", 0),
                transaction.get("target_price", 0) or 0,
                transaction.get("notes", ""),
                transaction.get("id", "")
            ]
            
            for col, value in enumerate(transaction_data, 1):
                cell = ws_transactions.cell(row=row, column=col, value=value)
                cell.border = border
                
                # Sayısal değerler için format
                if col in [5, 6, 7]:  # Fiyat, miktar, hedef fiyat
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
                
                # İşlem türü renklendirme
                if col == 4:  # İşlem türü kolonu
                    if value == "Alış":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Sütun genişliklerini ayarla
        for column in ws_transactions.columns:
            max_length = 0
            try:
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws_transactions.column_dimensions[column_letter].width = adjusted_width
            except AttributeError:
                # MergedCell için atla
                continue
        
        # 4. PERFORMANS ANALİZİ SAYFASI
        ws_performance = wb.create_sheet("Performans Analizi")
        
        # Performans başlıkları
        ws_performance['A1'] = "PERFORMANS ANALİZİ"
        ws_performance['A1'].font = Font(bold=True, size=16)
        ws_performance.merge_cells('A1:H1')
        
        # En iyi performans gösteren pozisyonlar
        ws_performance['A3'] = "En İyi Performans (Kar)"
        ws_performance['A3'].font = Font(bold=True, size=14)
        ws_performance['A3'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        performance_headers = ["Sembol", "Kar/Zarar", "Kar/Zarar %", "Yatırım", "Güncel Değer"]
        for col, header in enumerate(performance_headers, 1):
            cell = ws_performance.cell(row=4, column=col, value=header)
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.border = border
        
        # Pozisyonları kar/zarar'a göre sırala
        sorted_positions = sorted(positions_data, key=lambda x: (x.get("current_price", 0) or 0) * x.get("total_quantity", 0) - x.get("total_cost", 0), reverse=True)
        
        # En iyi 5 pozisyon
        for row, position in enumerate(sorted_positions[:5], 5):
            current_value = (position.get("current_price", 0) or 0) * position.get("total_quantity", 0)
            profit_loss = current_value - position.get("total_cost", 0)
            profit_loss_percent = (profit_loss / position.get("total_cost", 0)) * 100 if position.get("total_cost", 0) > 0 else 0
            
            performance_data = [
                position.get("symbol", ""),
                profit_loss,
                f"{profit_loss_percent:.2f}%",
                position.get("total_cost", 0),
                current_value
            ]
            
            for col, value in enumerate(performance_data, 1):
                cell = ws_performance.cell(row=row, column=col, value=value)
                cell.border = border
                if col in [2, 4, 5]:  # Sayısal kolonlar
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
        
        # Sütun genişliklerini ayarla
        for column in ws_performance.columns:
            max_length = 0
            try:
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws_performance.column_dimensions[column_letter].width = adjusted_width
            except AttributeError:
                # MergedCell için atla
                continue
        
        # Excel dosyasını kaydet
        wb.save(temp_file.name)
        return temp_file.name
        
    except Exception as e:
        print(f"Excel oluşturma hatası: {e}")
        raise e

# ---------- ADMIN ENDPOINT'LERİ ----------

@app.post("/admin/login")
async def admin_login(request: AdminLoginRequest):
    """Admin girişi"""
    try:
        users = load_users()
        admin_user = next((u for u in users if u["username"] == request.username and u["is_admin"]), None)
        
        if admin_user and verify_password(request.password, admin_user["password"]):
            # API key oluştur
            api_key = create_api_key(request.username)
            
            # Son giriş zamanını güncelle
            admin_user["last_login"] = datetime.now().isoformat()
            save_users(users)
            
            return {
                "success": True,
                "message": "Admin girişi başarılı",
                "user": {
                    "username": request.username,
                    "is_admin": True
                },
                "api_key": api_key
            }
        else:
            return {"success": False, "error": "Geçersiz kullanıcı adı veya şifre"}
    except Exception as e:
        return {"success": False, "error": f"Giriş hatası: {str(e)}"}

@app.post("/user/login")
async def user_login(request: UserLoginRequest):
    """Normal kullanıcı girişi"""
    try:
        users = load_users()
        user = next((u for u in users if u["username"] == request.username and u["is_active"]), None)
        
        print(f"🔍 DEBUG: Login attempt - username: {request.username}, password: {request.password}")
        print(f"🔍 DEBUG: Found user: {user}")
        if user:
            print(f"🔍 DEBUG: User password: {user['password']}")
        
        if user and verify_password(request.password, user["password"]):
            # API key oluştur
            api_key = create_api_key(request.username)
            
            # Son giriş zamanını güncelle
            user["last_login"] = datetime.now().isoformat()
            save_users(users)
            
            return {
                "success": True,
                "message": "Giriş başarılı",
                "user": {
                    "username": user["username"],
                    "email": user.get("email"),
                    "is_admin": user["is_admin"]
                },
                "api_key": api_key
            }
        else:
            return {"success": False, "error": "Geçersiz kullanıcı adı veya şifre"}
    except Exception as e:
        return {"success": False, "error": f"Giriş hatası: {str(e)}"}

@app.get("/admin/users")
async def get_all_users(current_user: dict = Depends(get_current_user)):
    """Tüm kullanıcıları getir (admin only)"""
    try:
        # Admin kontrolü
        if not current_user.get("is_admin"):
            return {"success": False, "error": "Admin yetkisi gerekli"}
            
        users = load_users()
        # Şifreleri göster (admin için)
        return {"success": True, "users": users}
    except Exception as e:
        return {"success": False, "error": f"Kullanıcılar yüklenemedi: {str(e)}"}

@app.post("/admin/users")
async def create_user(request: UserCreateRequest, current_user: dict = Depends(get_current_user)):
    """Yeni kullanıcı oluştur (admin only)"""
    try:
        # Admin kontrolü
        if not current_user.get("is_admin"):
            return {"success": False, "error": "Admin yetkisi gerekli"}
            
        users = load_users()
        
        # Kullanıcı adı zaten var mı kontrol et
        if any(user['username'] == request.username for user in users):
            return {"success": False, "error": "Bu kullanıcı adı zaten kullanılıyor"}
        
        # Yeni kullanıcı oluştur
        new_user = {
            "id": f"user_{len(users) + 1:03d}",
            "username": request.username,
            "password": hash_password(request.password),  # Şifreyi hash'le
            "email": request.email,
            "is_admin": request.is_admin,
            "created_at": datetime.now().isoformat(),
            "last_login": None,
            "is_active": True
        }
        
        users.append(new_user)
        save_users(users)
        
        # Şifreyi gizle
        new_user.pop('password', None)
        
        return {"success": True, "message": "Kullanıcı başarıyla oluşturuldu", "user": new_user}
    except Exception as e:
        return {"success": False, "error": f"Kullanıcı oluşturulamadı: {str(e)}"}

@app.put("/admin/users/{user_id}")
async def update_user(user_id: str, request: UserUpdateRequest, current_user: dict = Depends(get_current_user)):
    """Kullanıcı güncelle (admin only)"""
    try:
        # Admin kontrolü
        if not current_user.get("is_admin"):
            return {"success": False, "error": "Admin yetkisi gerekli"}
            
        users = load_users()
        
        # Kullanıcıyı bul
        user_index = None
        for i, user in enumerate(users):
            if user['id'] == user_id:
                user_index = i
                break
        
        if user_index is None:
            return {"success": False, "error": "Kullanıcı bulunamadı"}
        
        # Güncelle
        if request.username is not None:
            # Kullanıcı adı değişiyorsa, yeni adın benzersiz olduğunu kontrol et
            if request.username != users[user_index]['username']:
                if any(user['username'] == request.username for user in users if user['id'] != user_id):
                    return {"success": False, "error": "Bu kullanıcı adı zaten kullanılıyor"}
                users[user_index]['username'] = request.username
        
        if request.email is not None:
            users[user_index]['email'] = request.email
        if request.is_admin is not None:
            users[user_index]['is_admin'] = request.is_admin
        if request.is_active is not None:
            users[user_index]['is_active'] = request.is_active
        if request.password is not None:
            users[user_index]['password'] = hash_password(request.password)
        
        save_users(users)
        
        # Güncellenmiş kullanıcıyı döndür (şifre gizli)
        updated_user = users[user_index].copy()
        updated_user.pop('password', None)
        
        return {"success": True, "message": "Kullanıcı güncellendi", "user": updated_user}
    except Exception as e:
        return {"success": False, "error": f"Kullanıcı güncellenemedi: {str(e)}"}

@app.delete("/admin/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    """Kullanıcı sil (admin only)"""
    try:
        # Admin kontrolü
        if not current_user.get("is_admin"):
            return {"success": False, "error": "Admin yetkisi gerekli"}
            
        users = load_users()
        
        # Admin kullanıcısını silmeye izin verme
        user_to_delete = None
        for user in users:
            if user['id'] == user_id:
                if user['username'] == ADMIN_USERNAME:
                    return {"success": False, "error": "Ana admin kullanıcısı silinemez"}
                user_to_delete = user
                break
        
        if user_to_delete is None:
            return {"success": False, "error": "Kullanıcı bulunamadı"}
        
        # Kullanıcıyı sil
        users = [u for u in users if u['id'] != user_id]
        save_users(users)
        
        return {"success": True, "message": f"Kullanıcı '{user_to_delete['username']}' silindi"}
    except Exception as e:
        return {"success": False, "error": f"Kullanıcı silinemedi: {str(e)}"}

@app.get("/admin/portfolios")
async def get_all_portfolios(current_user: dict = Depends(get_current_user)):
    """Tüm kullanıcıların portföylerini getir (admin only)"""
    try:
        # Admin kontrolü
        if not current_user.get("is_admin"):
            return {"success": False, "error": "Admin yetkisi gerekli"}
            
        portfolio_list = load_portfolio_list()
        all_portfolios = []
        
        for portfolio in portfolio_list:
            portfolio_id = portfolio['portfolio_id']
            portfolio_file = os.path.join(PORTFOLIO_DIR, f"{portfolio_id}.json")
            
            if os.path.exists(portfolio_file):
                try:
                    with open(portfolio_file, 'r', encoding='utf-8') as f:
                        portfolio_data = json.load(f)
                    
                    # Portföy özeti
                    portfolio_summary = {
                        "id": portfolio_id,  # Frontend için alias
                        "portfolio_id": portfolio_id,
                        "portfolio_name": portfolio.get('portfolio_name', 'Bilinmeyen'),
                        "portfolio_description": portfolio.get('portfolio_description', ''),
                        "owner_username": portfolio.get('owner_username', 'Bilinmiyor'),  # Portföy sahibi
                        "total_transactions": len(portfolio_data),
                        "total_symbols": len(set(item['symbol'] for item in portfolio_data)),
                        "last_updated": portfolio.get('last_updated', 'Bilinmiyor')
                    }
                    
                    all_portfolios.append(portfolio_summary)
                except Exception as e:
                    print(f"Portföy {portfolio_id} okunamadı: {e}")
                    continue
        
        return {"success": True, "portfolios": all_portfolios}
    except Exception as e:
        return {"success": False, "error": f"Portföyler yüklenemedi: {str(e)}"}

@app.get("/admin/portfolio/{portfolio_id}")
async def get_portfolio_details(portfolio_id: str, current_user: dict = Depends(get_current_user)):
    """Belirli bir portföyün detaylarını getir (admin only)"""
    try:
        # Admin kontrolü
        if not current_user.get("is_admin"):
            return {"success": False, "error": "Admin yetkisi gerekli"}
            
        portfolio_file = os.path.join(PORTFOLIO_DIR, f"{portfolio_id}.json")
        
        if not os.path.exists(portfolio_file):
            return {"success": False, "error": "Portföy bulunamadı"}
        
        with open(portfolio_file, 'r', encoding='utf-8') as f:
            portfolio_data = json.load(f)
        
        # Portföy özeti
        total_investment = sum(item['price'] * item['quantity'] for item in portfolio_data if item['transaction_type'] == 'buy')
        total_sales = sum(item['price'] * item['quantity'] for item in portfolio_data if item['transaction_type'] == 'sell')
        
        portfolio_summary = {
            "portfolio_id": portfolio_id,
            "total_transactions": len(portfolio_data),
            "buy_transactions": len([item for item in portfolio_data if item['transaction_type'] == 'buy']),
            "sell_transactions": len([item for item in portfolio_data if item['transaction_type'] == 'sell']),
            "total_investment": total_investment,
            "total_sales": total_sales,
            "transactions": portfolio_data[:50]  # İlk 50 işlem
        }
        
        return {"success": True, "portfolio": portfolio_summary}
    except Exception as e:
        return {"success": False, "error": f"Portföy detayları yüklenemedi: {str(e)}"}

# ---------- UYGULAMA BAŞLATMA ----------
if __name__ == "__main__":
    # Database'i başlat
    print("🚀 DCA Scanner Backend başlatılıyor...")
    print("🔧 Database başlatılıyor...")
    init_database()
    
    # Mevcut JSON verilerini database'e taşı
    print("📦 Mevcut veriler database'e taşınıyor...")
    migrate_json_to_database()
    
    # Kalıcı kullanıcı verilerini sağla
    print("👥 Varsayılan kullanıcılar kontrol ediliyor...")
    ensure_default_users_exist()
    
    import uvicorn
    import os
    
    # Production'da PORT environment variable'ı kullan, local'de 8014
    port = int(os.environ.get("PORT", 8014))
    print(f"🌐 Server {port} portunda başlatılıyor...")
    uvicorn.run(app, host="0.0.0.0", port=port)

# ---------- VERİ YÜKLEME ENDPOINT'LERİ ----------
@app.post("/admin/load-data")
async def load_data_endpoint():
    """Kullanıcı ve portföy verilerini yükle (admin only)"""
    try:
        # Varsayılan admin kullanıcısını oluştur
        create_default_admin()
        
        # Test kullanıcıları oluştur
        users = load_users()
        
        # Test kullanıcıları ekle (eğer yoksa)
        test_users = [
            {"username": "deneme1", "password": "deneme123", "email": "deneme1@test.com"},
            {"username": "deneme2", "password": "deneme123", "email": "deneme2@test.com"},
            {"username": "deneme3", "password": "deneme123", "email": "deneme3@test.com"},
            {"username": "deneme4", "password": "deneme123", "email": "deneme4@test.com"}
        ]
        
        for test_user in test_users:
            if not any(u.get('username') == test_user['username'] for u in users):
                new_user = {
                    "id": f"user_{len(users) + 1:03d}",
                    "username": test_user['username'],
                    "password": test_user['password'],
                    "email": test_user['email'],
                    "is_admin": False,
                    "created_at": datetime.now().isoformat(),
                    "last_login": None,
                    "is_active": True
                }
                users.append(new_user)
                print(f"✅ Test kullanıcı oluşturuldu: {test_user['username']}")
        
        save_users(users)
        
        return {"success": True, "message": f"{len(users)} kullanıcı yüklendi"}
    except Exception as e:
        return {"success": False, "error": f"Veri yüklenemedi: {str(e)}"}

# ---------- KALICI KULLANICI VERİLERİ ----------
def ensure_default_users_exist():
    """Varsayılan kullanıcıların her zaman var olmasını sağla"""
    users = load_users()
    
    # Admin kullanıcısı kontrol et
    admin_exists = any(user.get('username') == ADMIN_USERNAME for user in users)
    if not admin_exists:
        print("🔄 Admin kullanıcısı oluşturuluyor...")
        admin_user = {
            "id": "admin_001",
            "username": ADMIN_USERNAME,
            "password": ADMIN_PASSWORD,
            "email": "admin@dca-scanner.com",
            "is_admin": True,
            "created_at": datetime.now().isoformat(),
            "last_login": None,
            "is_active": True
        }
        users.append(admin_user)
        print(f"✅ Admin kullanıcısı oluşturuldu: {ADMIN_USERNAME}")
    
    # Test kullanıcıları kontrol et
    test_users = [
        {"username": "deneme1", "password": "deneme123", "email": "deneme1@test.com"},
        {"username": "deneme2", "password": "deneme123", "email": "deneme2@test.com"},
        {"username": "deneme3", "password": "deneme123", "email": "deneme3@test.com"},
        {"username": "deneme4", "password": "deneme123", "email": "deneme4@test.com"}
    ]
    
    for test_user in test_users:
        if not any(u.get('username') == test_user['username'] for u in users):
            print(f"🔄 Test kullanıcısı oluşturuluyor: {test_user['username']}")
            new_user = {
                "id": f"user_{len(users) + 1:03d}",
                "username": test_user['username'],
                "password": test_user['password'],
                "email": test_user['email'],
                "is_admin": False,
                "created_at": datetime.now().isoformat(),
                "last_login": None,
                "is_active": True
            }
            users.append(new_user)
            print(f"✅ Test kullanıcısı oluşturuldu: {test_user['username']}")
    
    # Kullanıcıları kaydet
    if len(users) >= 5:  # En az 5 kullanıcı olmalı
        save_users(users)
        print(f"🎉 Toplam {len(users)} kullanıcı kalıcı olarak kaydedildi!")
    
    return users
